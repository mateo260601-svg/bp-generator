from openpyxl.utils import get_column_letter
from styles import *

FIRST_DATA_COL = 5
LABEL_COL = 2
UNIT_COL  = 3
SRC_COL   = 4

def _col(idx):
    return get_column_letter(FIRST_DATA_COL + idx)

def _data_row(ws, row, label, formulas, font, num_fmt="k0",
              fill_=None, unit="$k", indent=0, row_ht=16, border=None):
    ws.row_dimensions[row].height = row_ht
    prefix = "    " * indent
    c = ws.cell(row=row, column=LABEL_COL, value=prefix + label)
    style_cell(c, font=font, align=AL["left"])
    c2 = ws.cell(row=row, column=UNIT_COL, value=unit)
    style_cell(c2, font=FONTS["italic"], align=AL["center"])
    for i, f in enumerate(formulas):
        c = ws.cell(row=row, column=FIRST_DATA_COL + i, value=f)
        kw = dict(font=font, align=AL["right"], num_fmt=NF[num_fmt])
        if fill_:  kw["fill_"] = fill_
        if border: kw["border"] = border
        style_cell(c, **kw)

def _header_block(ws, row, title, color_key, n, row_ht=20):
    ws.row_dimensions[row].height = row_ht
    end = get_column_letter(FIRST_DATA_COL + n - 1)
    merge_and_style(ws, f"B{row}:{end}{row}", title,
                    FONTS["section"], fill(C[color_key]), AL["left"])
    return row + 1

def _period_headers(ws, row, years):
    n = len(years)
    ws.row_dimensions[row].height = 18
    for i, yr in enumerate(years):
        c = ws.cell(row=row, column=FIRST_DATA_COL + i, value=yr["fy"])
        style_cell(c, font=FONTS["header"], fill_=fill(C["navy"]), align=AL["center"])
    row += 1
    ws.row_dimensions[row].height = 14
    for i, yr in enumerate(years):
        lbl = "Actuals" if yr["is_actual"] else "BP"
        c = ws.cell(row=row, column=FIRST_DATA_COL + i, value=lbl)
        clr = C["mid_grey"] if yr["is_actual"] else C["pale_blue"]
        style_cell(c, font=FONTS["note"], fill_=fill(clr), align=AL["center"])
    return row + 1

def _sheet_setup(wb, name, tab_color, zoom=90):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = zoom
    ws.sheet_properties.tabColor = tab_color
    set_col_width(ws, "A", 2)
    set_col_width(ws, "B", 38)
    set_col_width(ws, "C", 8)
    set_col_width(ws, "D", 10)
    return ws

def _col_widths(ws, n):
    for i in range(n):
        set_col_width(ws, get_column_letter(FIRST_DATA_COL + i), 14)
    set_col_width(ws, get_column_letter(FIRST_DATA_COL + n), 28)

# ═══════════════════════════════════════════════════════════════════════════════
# NWC
# ═══════════════════════════════════════════════════════════════════════════════
def build_nwc(wb, config, years, row_map):
    ws = _sheet_setup(wb, "NWC", "7B3F00")
    n = len(years)
    _col_widths(ws, n)
    ccy = config["currency"]
    unit = f"{ccy}k"
    row = 1

    ws.row_dimensions[row].height = 34
    end = get_column_letter(FIRST_DATA_COL + n - 1)
    merge_and_style(ws, f"B{row}:{end}{row}",
                    f"{config['company_name']} — Net Working Capital",
                    FONTS["title"], fill(C["navy"]), AL["center"])
    row += 2
    row = _period_headers(ws, row, years)

    row = _header_block(ws, row, "WORKING CAPITAL COMPONENTS", "nwc_header", n)

    # DSO/DIO/DPO from ASSUMPTIONS
    dso_row = 8   # calibrated ASSUMPTIONS rows for NWC
    dio_row = 9
    dpo_row = 10
    rev_row = row_map.get("pl_net_real", 20)
    cogs_row = row_map.get("pl_cogs_var", 30)

    dso_fmls = [f"=ASSUMPTIONS!{_col(i)}{dso_row}" for i in range(n)]
    dio_fmls = [f"=ASSUMPTIONS!{_col(i)}{dio_row}" for i in range(n)]
    dpo_fmls = [f"=ASSUMPTIONS!{_col(i)}{dpo_row}" for i in range(n)]

    _data_row(ws, row, "Days Sales Outstanding (DSO)", dso_fmls,
              FONTS["link"], unit="days", num_fmt="int0")
    dso_r = row; row += 1

    _data_row(ws, row, "Days Inventory Outstanding (DIO)", dio_fmls,
              FONTS["link"], unit="days", num_fmt="int0")
    dio_r = row; row += 1

    _data_row(ws, row, "Days Payables Outstanding (DPO)", dpo_fmls,
              FONTS["link"], unit="days", num_fmt="int0")
    dpo_r = row; row += 2

    row = _header_block(ws, row, "WORKING CAPITAL BALANCES", "nwc_header", n)

    # Trade receivables = Net revenue × DSO / 365
    ar_fmls = [f"='P&L'!{_col(i)}{rev_row}*{_col(i)}{dso_r}/365" for i in range(n)]
    _data_row(ws, row, "Trade receivables (AR)", ar_fmls, FONTS["formula"], unit=unit)
    ar_row = row; row += 1

    inv_fmls = [f"='P&L'!{_col(i)}{cogs_row}*{_col(i)}{dio_r}/365" for i in range(n)]
    _data_row(ws, row, "Inventories", inv_fmls, FONTS["formula"], unit=unit)
    inv_row = row; row += 1

    oca_fmls = [f"='P&L'!{_col(i)}{rev_row}*ASSUMPTIONS!{_col(i)}11" for i in range(n)]
    _data_row(ws, row, "Other current assets", oca_fmls, FONTS["formula"], unit=unit)
    oca_row = row; row += 1

    ca_fmls = [f"={_col(i)}{ar_row}+{_col(i)}{inv_row}+{_col(i)}{oca_row}" for i in range(n)]
    _data_row(ws, row, "Total current assets", ca_fmls, FONTS["subtotal"],
              fill_=fill(C["subtot_fill"]), unit=unit, border=BORDERS["thin"])
    ca_row = row; row += 1

    ap_fmls = [f"='P&L'!{_col(i)}{cogs_row}*{_col(i)}{dpo_r}/365" for i in range(n)]
    _data_row(ws, row, "Trade payables (AP)", ap_fmls, FONTS["formula"], unit=unit)
    ap_row = row; row += 1

    ocl_fmls = [f"='P&L'!{_col(i)}{rev_row}*ASSUMPTIONS!{_col(i)}12" for i in range(n)]
    _data_row(ws, row, "Other current liabilities", ocl_fmls, FONTS["formula"], unit=unit)
    ocl_row = row; row += 1

    cl_fmls = [f"={_col(i)}{ap_row}+{_col(i)}{ocl_row}" for i in range(n)]
    _data_row(ws, row, "Total current liabilities", cl_fmls, FONTS["subtotal"],
              fill_=fill(C["subtot_fill"]), unit=unit, border=BORDERS["thin"])
    cl_row = row; row += 1

    nwc_fmls = [f"={_col(i)}{ca_row}-{_col(i)}{cl_row}" for i in range(n)]
    _data_row(ws, row, "Net working capital", nwc_fmls, FONTS["total"],
              fill_=fill(C["total_fill"]), unit=unit, border=BORDERS["thick"], row_ht=20)
    row_map["nwc_balance"] = row

    nwc_chg = [f"=IFERROR({_col(i)}{row}-{_col(i-1)}{row},0)" if i > 0 else "=0"
               for i in range(n)]
    row += 1
    _data_row(ws, row, "Change in NWC (increase = cash outflow)", nwc_chg,
              FONTS["formula"], unit=unit, indent=1)
    row_map["nwc_change"] = row

    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# CAPEX
# ═══════════════════════════════════════════════════════════════════════════════
def build_capex(wb, config, years, row_map):
    ws = _sheet_setup(wb, "CAPEX", "2C4A6E")
    n = len(years)
    _col_widths(ws, n)
    ccy = config["currency"]
    unit = f"{ccy}k"
    cpx = config.get("capex", {})
    ul = cpx.get("useful_life", 20)
    open_ppe = cpx.get("opening_ppe", 85000)
    open_dep = cpx.get("accum_dep_open", 24000)
    row = 1

    ws.row_dimensions[row].height = 34
    end = get_column_letter(FIRST_DATA_COL + n - 1)
    merge_and_style(ws, f"B{row}:{end}{row}",
                    f"{config['company_name']} — CAPEX & Depreciation Schedule",
                    FONTS["title"], fill(C["navy"]), AL["center"])
    row += 2
    row = _period_headers(ws, row, years)

    row = _header_block(ws, row, "CAPITAL EXPENDITURE", "kpi_header", n)

    maint_fmls = [f"=ASSUMPTIONS!{_col(i)}22" for i in range(n)]
    expan_fmls = [f"=ASSUMPTIONS!{_col(i)}23" for i in range(n)]
    _data_row(ws, row, "Maintenance capex", maint_fmls, FONTS["link"], unit=unit)
    maint_row = row; row += 1
    _data_row(ws, row, "Expansion capex",   expan_fmls, FONTS["link"], unit=unit)
    expan_row = row; row += 1

    tot_cpx = [f"={_col(i)}{maint_row}+{_col(i)}{expan_row}" for i in range(n)]
    _data_row(ws, row, "Total capex", tot_cpx, FONTS["total"],
              fill_=fill(C["total_fill"]), unit=unit, border=BORDERS["thick"])
    row_map["capex_total"] = row
    cpx_tot_row = row; row += 2

    row = _header_block(ws, row, "PP&E ROLL-FORWARD (GROSS)", "kpi_header", n)

    ppe_open = [None] * n
    ppe_open[0] = open_ppe
    for i in range(1, n):
        ppe_open[i] = f"={_col(i-1)}{row+2}"  # will point to closing row
    _data_row(ws, row, "Opening PP&E (gross)", ppe_open, FONTS["input"],
              fill_=fill(C["input_bg"]), unit=unit)
    ppe_open_row = row; row += 1

    add_fmls = [f"={_col(i)}{cpx_tot_row}" for i in range(n)]
    _data_row(ws, row, "Additions (capex)", add_fmls, FONTS["formula"], unit=unit)
    ppe_add_row = row; row += 1

    disp_fmls = [f"=0" for _ in range(n)]  # disposals placeholder
    _data_row(ws, row, "Disposals", disp_fmls, FONTS["input"],
              fill_=fill(C["input_bg"]), unit=unit)
    ppe_disp_row = row; row += 1

    close_fmls = [f"={_col(i)}{ppe_open_row}+{_col(i)}{ppe_add_row}-{_col(i)}{ppe_disp_row}"
                  for i in range(n)]
    _data_row(ws, row, "Closing PP&E (gross)", close_fmls, FONTS["total"],
              fill_=fill(C["total_fill"]), unit=unit, border=BORDERS["thick"])
    ppe_close_row = row; row += 2

    # Fix forward-references for opening PP&E years 2+
    for i in range(1, n):
        ws.cell(row=ppe_open_row, column=FIRST_DATA_COL + i).value = \
            f"={_col(i-1)}{ppe_close_row}"

    row = _header_block(ws, row, "ACCUMULATED DEPRECIATION", "kpi_header", n)

    dep_ann = [f"={_col(i)}{ppe_open_row}/{ul}" for i in range(n)]
    _data_row(ws, row, "Annual depreciation charge", dep_ann, FONTS["formula"], unit=unit)
    dep_ann_row = row
    row_map["capex_dep_row"] = row  # referenced by P&L as row 20 equivalent
    row += 1

    accum_open = [None] * n
    accum_open[0] = open_dep
    _data_row(ws, row, "Opening accumulated depreciation", accum_open,
              FONTS["input"], fill_=fill(C["input_bg"]), unit=unit)
    acc_open_row = row; row += 1

    accum_close_fmls = [f"={_col(i)}{acc_open_row}+{_col(i)}{dep_ann_row}"
                        for i in range(n)]
    _data_row(ws, row, "Closing accumulated depreciation", accum_close_fmls,
              FONTS["formula"], unit=unit)
    acc_close_row = row; row += 2

    for i in range(1, n):
        ws.cell(row=acc_open_row, column=FIRST_DATA_COL + i).value = \
            f"={_col(i-1)}{acc_close_row}"

    row = _header_block(ws, row, "NET BOOK VALUE (NBV)", "kpi_header", n)
    nbv_fmls = [f"={_col(i)}{ppe_close_row}-{_col(i)}{acc_close_row}" for i in range(n)]
    _data_row(ws, row, "Net PP&E (NBV)", nbv_fmls, FONTS["total"],
              fill_=fill(C["total_fill"]), unit=unit, border=BORDERS["thick"])
    row_map["capex_nbv"] = row

    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# BALANCE SHEET
# ═══════════════════════════════════════════════════════════════════════════════
def build_balance_sheet(wb, config, years, row_map):
    ws = _sheet_setup(wb, "BALANCE SHEET", "1E4D2B")
    n = len(years)
    _col_widths(ws, n)
    ccy = config["currency"]
    unit = f"{ccy}k"
    row = 1

    ws.row_dimensions[row].height = 34
    end = get_column_letter(FIRST_DATA_COL + n - 1)
    merge_and_style(ws, f"B{row}:{end}{row}",
                    f"{config['company_name']} — Balance Sheet",
                    FONTS["title"], fill(C["navy"]), AL["center"])
    row += 2
    row = _period_headers(ws, row, years)

    # ── ASSETS ──────────────────────────────────────────────────────────────
    row = _header_block(ws, row, "ASSETS", "bs_header", n)

    row = _header_block(ws, row, "Non-current assets", "bs_sub", n, row_ht=17)
    nbv_fmls = [f"=IFERROR(CAPEX!{_col(i)}{row_map.get('capex_nbv',30)},0)" for i in range(n)]
    _data_row(ws, row, "Property, plant & equipment (net)", nbv_fmls, FONTS["link"], unit=unit)
    ppe_row = row; row += 1

    goodwill_fmls = [f"=0" for _ in range(n)]
    _data_row(ws, row, "Goodwill & intangibles", goodwill_fmls, FONTS["input"],
              fill_=fill(C["input_bg"]), unit=unit)
    gw_row = row; row += 1

    other_nca = [f"=0" for _ in range(n)]
    _data_row(ws, row, "Other non-current assets", other_nca, FONTS["input"],
              fill_=fill(C["input_bg"]), unit=unit)
    onca_row = row; row += 1

    nca_fmls = [f"={_col(i)}{ppe_row}+{_col(i)}{gw_row}+{_col(i)}{onca_row}" for i in range(n)]
    _data_row(ws, row, "Total non-current assets", nca_fmls, FONTS["subtotal"],
              fill_=fill(C["subtot_fill"]), unit=unit, border=BORDERS["thin"])
    nca_tot_row = row; row += 1

    row = _header_block(ws, row, "Current assets", "bs_sub", n, row_ht=17)
    ar_fmls   = [f"=IFERROR(NWC!{_col(i)}{row_map.get('nwc_balance',20)-2},0)" for i in range(n)]
    inv_fmls  = [f"=IFERROR(NWC!{_col(i)}{row_map.get('nwc_balance',20)-1},0)" for i in range(n)]
    _data_row(ws, row, "Trade receivables", ar_fmls, FONTS["link"], unit=unit)
    ar_row = row; row += 1
    _data_row(ws, row, "Inventories", inv_fmls, FONTS["link"], unit=unit)
    inv_row = row; row += 1

    cash_fmls = [f"=IFERROR('CASH FLOW'!{_col(i)}{row_map.get('cf_closing_cash',50)},0)"
                 for i in range(n)]
    _data_row(ws, row, "Cash & cash equivalents", cash_fmls, FONTS["link"], unit=unit)
    cash_row = row; row += 1

    oca_fmls = [f"=0" for _ in range(n)]
    _data_row(ws, row, "Other current assets", oca_fmls, FONTS["input"],
              fill_=fill(C["input_bg"]), unit=unit)
    oca_row = row; row += 1

    ca_fmls = [f"=SUM({_col(i)}{ar_row}:{_col(i)}{oca_row})" for i in range(n)]
    _data_row(ws, row, "Total current assets", ca_fmls, FONTS["subtotal"],
              fill_=fill(C["subtot_fill"]), unit=unit, border=BORDERS["thin"])
    ca_tot_row = row; row += 1

    total_assets_fmls = [f"={_col(i)}{nca_tot_row}+{_col(i)}{ca_tot_row}" for i in range(n)]
    _data_row(ws, row, "TOTAL ASSETS", total_assets_fmls, FONTS["total"],
              fill_=fill(C["total_fill"]), unit=unit, border=BORDERS["thick"], row_ht=22)
    row_map["bs_total_assets"] = row
    ta_row = row; row += 2

    # ── LIABILITIES ──────────────────────────────────────────────────────────
    row = _header_block(ws, row, "LIABILITIES", "bs_header", n)

    row = _header_block(ws, row, "Non-current liabilities", "bs_sub", n, row_ht=17)
    debt_fmls = [f"=IFERROR('DEBT SCHEDULE'!{_col(i)}{row_map.get('ds_total_debt',10)},0)"
                 for i in range(n)]
    _data_row(ws, row, "Financial debt (non-current)", debt_fmls, FONTS["link"], unit=unit)
    ltd_row = row; row += 1
    dt_fmls = [f"=0" for _ in range(n)]
    _data_row(ws, row, "Deferred tax liabilities", dt_fmls, FONTS["formula"], unit=unit)
    dtl_row = row; row += 1
    onl_fmls = [f"=0" for _ in range(n)]
    _data_row(ws, row, "Other non-current liabilities", onl_fmls, FONTS["input"],
              fill_=fill(C["input_bg"]), unit=unit)
    onl_row = row; row += 1

    ncl_fmls = [f"=SUM({_col(i)}{ltd_row}:{_col(i)}{onl_row})" for i in range(n)]
    _data_row(ws, row, "Total non-current liabilities", ncl_fmls, FONTS["subtotal"],
              fill_=fill(C["subtot_fill"]), unit=unit, border=BORDERS["thin"])
    ncl_tot_row = row; row += 1

    row = _header_block(ws, row, "Current liabilities", "bs_sub", n, row_ht=17)
    ap_fmls  = [f"=IFERROR(NWC!{_col(i)}{row_map.get('nwc_balance',20)+1},0)" for i in range(n)]
    _data_row(ws, row, "Trade payables", ap_fmls, FONTS["link"], unit=unit)
    ap_row = row; row += 1
    stdebt_fmls = [f"=0" for _ in range(n)]
    _data_row(ws, row, "Current portion of debt", stdebt_fmls, FONTS["formula"], unit=unit)
    std_row = row; row += 1
    ocl_fmls = [f"=0" for _ in range(n)]
    _data_row(ws, row, "Other current liabilities", ocl_fmls, FONTS["input"],
              fill_=fill(C["input_bg"]), unit=unit)
    ocl_row = row; row += 1

    cl_fmls = [f"=SUM({_col(i)}{ap_row}:{_col(i)}{ocl_row})" for i in range(n)]
    _data_row(ws, row, "Total current liabilities", cl_fmls, FONTS["subtotal"],
              fill_=fill(C["subtot_fill"]), unit=unit, border=BORDERS["thin"])
    cl_tot_row = row; row += 1

    total_liab_fmls = [f"={_col(i)}{ncl_tot_row}+{_col(i)}{cl_tot_row}" for i in range(n)]
    _data_row(ws, row, "TOTAL LIABILITIES", total_liab_fmls, FONTS["total"],
              fill_=fill(C["total_fill"]), unit=unit, border=BORDERS["thick"], row_ht=22)
    row_map["bs_total_liab"] = row
    tl_row = row; row += 2

    # ── EQUITY ───────────────────────────────────────────────────────────────
    row = _header_block(ws, row, "EQUITY", "bs_header", n)
    sc_fmls = [f"=0" for _ in range(n)]
    _data_row(ws, row, "Share capital", sc_fmls, FONTS["input"],
              fill_=fill(C["input_bg"]), unit=unit)
    sc_row = row; row += 1
    re_fmls = [f"=IFERROR('P&L'!{_col(i)}{row_map.get('pl_net_income',80)},0)" for i in range(n)]
    _data_row(ws, row, "Retained earnings (current year)", re_fmls, FONTS["link"], unit=unit)
    re_row = row; row += 1
    accre_fmls = [f"=IFERROR({_col(i-1)}{re_row}+{_col(i-1)}{row+0},0)" if i > 0 else "=0"
                  for i in range(n)]
    _data_row(ws, row, "Accumulated retained earnings", accre_fmls, FONTS["formula"], unit=unit)
    accre_row = row; row += 1

    teq_fmls = [f"=SUM({_col(i)}{sc_row}:{_col(i)}{accre_row})" for i in range(n)]
    _data_row(ws, row, "TOTAL EQUITY", teq_fmls, FONTS["total"],
              fill_=fill(C["total_fill"]), unit=unit, border=BORDERS["thick"], row_ht=22)
    row_map["bs_total_equity"] = row
    teq_row = row; row += 1

    tl_eq_fmls = [f"={_col(i)}{tl_row}+{_col(i)}{teq_row}" for i in range(n)]
    _data_row(ws, row, "TOTAL LIABILITIES & EQUITY", tl_eq_fmls, FONTS["total"],
              fill_=fill(C["navy"]), unit=unit, border=BORDERS["thick"], row_ht=22)
    row_map["bs_total_l_and_e"] = row
    tle_row = row; row += 2

    # ── CHECK ────────────────────────────────────────────────────────────────
    row = _header_block(ws, row, "BALANCE SHEET CHECK", "kpi_header", n, row_ht=18)
    check_fmls = [
        f"=IF(ABS({_col(i)}{ta_row}-{_col(i)}{tle_row})<1,\"✓ Balanced\",\"✗ ERROR: \"&TEXT(ABS({_col(i)}{ta_row}-{_col(i)}{tle_row}),\"#,##0\"))"
        for i in range(n)]
    for i in range(n):
        c = ws.cell(row=row, column=FIRST_DATA_COL + i, value=check_fmls[i])
        style_cell(c, font=FONTS["formula"], align=AL["center"])
    c_lbl = ws.cell(row=row, column=LABEL_COL, value="Assets = Liabilities + Equity")
    style_cell(c_lbl, font=FONTS["label_bold"], align=AL["left"])

    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# CASH FLOW
# ═══════════════════════════════════════════════════════════════════════════════
def build_cash_flow(wb, config, years, row_map):
    ws = _sheet_setup(wb, "CASH FLOW", "4A235A")
    n = len(years)
    _col_widths(ws, n)
    ccy = config["currency"]
    unit = f"{ccy}k"
    row = 1

    ws.row_dimensions[row].height = 34
    end = get_column_letter(FIRST_DATA_COL + n - 1)
    merge_and_style(ws, f"B{row}:{end}{row}",
                    f"{config['company_name']} — Cash Flow Statement",
                    FONTS["title"], fill(C["navy"]), AL["center"])
    row += 2
    row = _period_headers(ws, row, years)

    def pl(key): return row_map.get(key, 5)

    # ── OPERATING ──────────────────────────────────────────────────────────
    row = _header_block(ws, row, "A. OPERATING CASH FLOW", "cf_header", n)

    ebitda_fmls = [f"='P&L'!{_col(i)}{pl('pl_ebitda_adj')}" for i in range(n)]
    _data_row(ws, row, "Adjusted EBITDA", ebitda_fmls, FONTS["link"], unit=unit)
    ebitda_r = row; row += 1

    nwc_chg_fmls = [f"=-IFERROR(NWC!{_col(i)}{row_map.get('nwc_change',25)},0)" for i in range(n)]
    _data_row(ws, row, "Change in net working capital", nwc_chg_fmls,
              FONTS["link"], unit=unit, indent=1)
    nwc_r = row; row += 1

    tax_fmls = [f"=-IFERROR('P&L'!{_col(i)}{pl('pl_curr_tax')},0)" for i in range(n)]
    _data_row(ws, row, "Tax paid (current)", tax_fmls, FONTS["link"], unit=unit, indent=1)
    tax_r = row; row += 1

    others_r_fmls = [f"=0" for _ in range(n)]
    _data_row(ws, row, "Other operating items", others_r_fmls, FONTS["input"],
              fill_=fill(C["input_bg"]), unit=unit, indent=1)
    other_op_r = row; row += 1

    op_cf_fmls = [f"={_col(i)}{ebitda_r}+{_col(i)}{nwc_r}+{_col(i)}{tax_r}+{_col(i)}{other_op_r}"
                  for i in range(n)]
    _data_row(ws, row, "Cash flow from operations", op_cf_fmls, FONTS["total"],
              fill_=fill(C["total_fill"]), unit=unit, border=BORDERS["thick"], row_ht=20)
    row_map["cf_operating"] = row
    op_row = row; row += 2

    # ── INVESTING ─────────────────────────────────────────────────────────
    row = _header_block(ws, row, "B. INVESTING CASH FLOW", "cf_header", n)

    cpx_fmls = [f"=-IFERROR(CAPEX!{_col(i)}{row_map.get('capex_total',20)},0)" for i in range(n)]
    _data_row(ws, row, "Capital expenditure", cpx_fmls, FONTS["link"], unit=unit, indent=1)
    cpx_r = row; row += 1
    disp_fmls = [f"=0" for _ in range(n)]
    _data_row(ws, row, "Proceeds from disposals", disp_fmls, FONTS["input"],
              fill_=fill(C["input_bg"]), unit=unit, indent=1)
    disp_r = row; row += 1
    acq_fmls = [f"=0" for _ in range(n)]
    _data_row(ws, row, "Acquisitions / investments", acq_fmls, FONTS["input"],
              fill_=fill(C["input_bg"]), unit=unit, indent=1)
    acq_r = row; row += 1

    inv_cf_fmls = [f"={_col(i)}{cpx_r}+{_col(i)}{disp_r}+{_col(i)}{acq_r}" for i in range(n)]
    _data_row(ws, row, "Cash flow from investing", inv_cf_fmls, FONTS["total"],
              fill_=fill(C["total_fill"]), unit=unit, border=BORDERS["thick"], row_ht=20)
    row_map["cf_investing"] = row
    inv_row = row; row += 2

    # ── FREE CASH FLOW ─────────────────────────────────────────────────────
    fcf_fmls = [f"={_col(i)}{op_row}+{_col(i)}{inv_row}" for i in range(n)]
    _data_row(ws, row, "FREE CASH FLOW (pre-debt service)", fcf_fmls, FONTS["total"],
              fill_=fill(C["ebitda_fill"]), unit=unit, border=BORDERS["thick"], row_ht=22)
    row_map["cf_fcf"] = row
    # Row 45 referenced from Debt Schedule for cash sweep
    fcf_row = row; row += 2

    # ── FINANCING ─────────────────────────────────────────────────────────
    row = _header_block(ws, row, "C. FINANCING CASH FLOW", "cf_header", n)

    int_fmls = [f"=-IFERROR('DEBT SCHEDULE'!{_col(i)}{row_map.get('ds_cash_interest',20)},0)"
                for i in range(n)]
    _data_row(ws, row, "Interest paid (cash)", int_fmls, FONTS["link"], unit=unit, indent=1)
    int_r = row; row += 1

    rep_fmls = [f"=-IFERROR('DEBT SCHEDULE'!{_col(i)}{row_map.get('ds_total_repay',15)},0)"
                for i in range(n)]
    _data_row(ws, row, "Debt repayments", rep_fmls, FONTS["link"], unit=unit, indent=1)
    rep_r = row; row += 1

    draw_fmls = [f"=0" for _ in range(n)]
    _data_row(ws, row, "New drawdowns", draw_fmls, FONTS["input"],
              fill_=fill(C["input_bg"]), unit=unit, indent=1)
    draw_r = row; row += 1

    div_fmls = [f"=0" for _ in range(n)]
    _data_row(ws, row, "Dividends / equity distributions", div_fmls, FONTS["input"],
              fill_=fill(C["input_bg"]), unit=unit, indent=1)
    div_r = row; row += 1

    fin_cf_fmls = [f"={_col(i)}{int_r}+{_col(i)}{rep_r}+{_col(i)}{draw_r}+{_col(i)}{div_r}"
                   for i in range(n)]
    _data_row(ws, row, "Cash flow from financing", fin_cf_fmls, FONTS["total"],
              fill_=fill(C["total_fill"]), unit=unit, border=BORDERS["thick"], row_ht=20)
    row_map["cf_financing"] = row
    fin_row = row; row += 2

    # ── NET MOVEMENT & CLOSING CASH ───────────────────────────────────────
    net_fmls = [f"={_col(i)}{op_row}+{_col(i)}{inv_row}+{_col(i)}{fin_row}" for i in range(n)]
    _data_row(ws, row, "Net movement in cash", net_fmls, FONTS["subtotal"],
              fill_=fill(C["subtot_fill"]), unit=unit, border=BORDERS["thin"])
    net_r = row; row += 1

    open_cash_fmls = [f"=IFERROR({_col(i-1)}{row+1},0)" if i > 0 else f"=0"
                      for i in range(n)]
    _data_row(ws, row, "Opening cash", open_cash_fmls, FONTS["input"],
              fill_=fill(C["input_bg"]), unit=unit)
    open_cash_r = row; row += 1

    close_fmls = [f"={_col(i)}{open_cash_r}+{_col(i)}{net_r}" for i in range(n)]
    _data_row(ws, row, "Closing cash", close_fmls, FONTS["total"],
              fill_=fill(C["total_fill"]), unit=unit, border=BORDERS["thick"], row_ht=22)
    row_map["cf_closing_cash"] = row

    # Fix forward refs for opening cash
    for i in range(1, n):
        ws.cell(row=open_cash_r, column=FIRST_DATA_COL + i).value = \
            f"={_col(i-1)}{row}"

    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# KPIs & RATIOS
# ═══════════════════════════════════════════════════════════════════════════════
def build_kpis(wb, config, years, row_map):
    ws = _sheet_setup(wb, "KPIs & RATIOS", "2C4A6E")
    n = len(years)
    _col_widths(ws, n)
    ccy = config["currency"]
    unit = f"{ccy}k"
    row = 1

    ws.row_dimensions[row].height = 34
    end = get_column_letter(FIRST_DATA_COL + n - 1)
    merge_and_style(ws, f"B{row}:{end}{row}",
                    f"{config['company_name']} — KPIs & Financial Ratios",
                    FONTS["title"], fill(C["navy"]), AL["center"])
    row += 2
    row = _period_headers(ws, row, years)

    def pl(k):  return row_map.get(k, 5)
    def ds(k):  return row_map.get(k, 5)

    # ── PROFITABILITY ─────────────────────────────────────────────────────
    row = _header_block(ws, row, "1. PROFITABILITY", "kpi_header", n)

    kpis_prof = [
        ("Revenue",          f"='P&L'!{{c}}{pl('pl_net_real')}",      unit,  "k0"),
        ("Gross profit",     f"='P&L'!{{c}}{pl('pl_gross_profit')}",   unit,  "k0"),
        ("Gross margin %",   f"=IFERROR('P&L'!{{c}}{pl('pl_gross_profit')}/'P&L'!{{c}}{pl('pl_net_real')},\"-\")", "%","pct1"),
        ("EBITDA (adj.)",    f"='P&L'!{{c}}{pl('pl_ebitda_adj')}",     unit,  "k0"),
        ("EBITDA margin %",  f"=IFERROR('P&L'!{{c}}{pl('pl_ebitda_adj')}/'P&L'!{{c}}{pl('pl_net_real')},\"-\")", "%","pct1"),
        ("EBIT",             f"='P&L'!{{c}}{pl('pl_ebit')}",           unit,  "k0"),
        ("EBIT margin %",    f"=IFERROR('P&L'!{{c}}{pl('pl_ebit')}/'P&L'!{{c}}{pl('pl_net_real')},\"-\")", "%","pct1"),
        ("Net income",       f"='P&L'!{{c}}{pl('pl_net_income')}",     unit,  "k0"),
        ("Net margin %",     f"=IFERROR('P&L'!{{c}}{pl('pl_net_income')}/'P&L'!{{c}}{pl('pl_net_real')},\"-\")", "%","pct1"),
    ]

    for label, fml_tpl, u, fmt in kpis_prof:
        fmls = [fml_tpl.replace("{c}", _col(i)) for i in range(n)]
        _data_row(ws, row, label, fmls, FONTS["formula"], unit=u, num_fmt=fmt)
        row += 1
    row += 1

    # ── LEVERAGE ─────────────────────────────────────────────────────────
    row = _header_block(ws, row, "2. LEVERAGE & COVERAGE", "kpi_header", n)

    gross_debt_fmls = [f"=IFERROR('DEBT SCHEDULE'!{_col(i)}{ds('ds_total_debt')},0)" for i in range(n)]
    _data_row(ws, row, "Gross debt", gross_debt_fmls, FONTS["link"], unit=unit)
    gd_row = row; row += 1

    cash_fmls = [f"=IFERROR('CASH FLOW'!{_col(i)}{pl('cf_closing_cash')},0)" for i in range(n)]
    _data_row(ws, row, "Cash & equivalents", cash_fmls, FONTS["link"], unit=unit)
    cash_r = row; row += 1

    nd_fmls = [f"={_col(i)}{gd_row}-{_col(i)}{cash_r}" for i in range(n)]
    _data_row(ws, row, "Net debt", nd_fmls, FONTS["total"],
              fill_=fill(C["total_fill"]), unit=unit, border=BORDERS["thick"])
    nd_row = row; row += 1

    lev_fmls = [f"=IFERROR({_col(i)}{nd_row}/'P&L'!{_col(i)}{pl('pl_ebitda_adj')},\"-\")"
                for i in range(n)]
    _data_row(ws, row, "Net leverage (ND / Adj EBITDA)", lev_fmls, FONTS["formula"],
              unit="x", num_fmt="mult")
    row += 1

    icr_fmls = [f"=IFERROR('P&L'!{_col(i)}{pl('pl_ebitda_adj')}/'P&L'!{_col(i)}{pl('pl_net_interest')},\"-\")"
                for i in range(n)]
    _data_row(ws, row, "Interest cover (Adj EBITDA / Net int.)", icr_fmls, FONTS["formula"],
              unit="x", num_fmt="mult")
    row += 1

    dscr_fmls = [f"=IFERROR('CASH FLOW'!{_col(i)}{pl('cf_operating')}/"
                 f"('P&L'!{_col(i)}{pl('pl_net_interest')}+'DEBT SCHEDULE'!{_col(i)}{ds('ds_total_repay')}),\"-\")"
                 for i in range(n)]
    _data_row(ws, row, "DSCR (Op.CF / (Interest + Repayment))", dscr_fmls, FONTS["formula"],
              unit="x", num_fmt="mult")
    row += 2

    # ── EFFICIENCY ────────────────────────────────────────────────────────
    row = _header_block(ws, row, "3. EFFICIENCY & RETURNS", "kpi_header", n)

    roce_fmls = [f"=IFERROR('P&L'!{_col(i)}{pl('pl_ebit')}/"
                 f"('BALANCE SHEET'!{_col(i)}{row_map.get('bs_total_assets',30)}-"
                 f"'BALANCE SHEET'!{_col(i)}{row_map.get('bs_total_liab',40)}+{_col(i)}{nd_row}),\"-\")"
                 for i in range(n)]
    _data_row(ws, row, "ROCE (EBIT / Capital employed)", roce_fmls, FONTS["formula"],
              unit="%", num_fmt="pct1")
    row += 1

    fcf_fmls = [f"=IFERROR('CASH FLOW'!{_col(i)}{pl('cf_fcf')},0)" for i in range(n)]
    _data_row(ws, row, "Free cash flow", fcf_fmls, FONTS["link"], unit=unit)
    row += 1

    capex_int_fmls = [f"=IFERROR(CAPEX!{_col(i)}{row_map.get('capex_total',20)}/"
                      f"'P&L'!{_col(i)}{pl('pl_ebitda_adj')},\"-\")"
                      for i in range(n)]
    _data_row(ws, row, "Capex intensity (Capex / Adj EBITDA)", capex_int_fmls,
              FONTS["formula"], unit="%", num_fmt="pct1")
    row += 2

    # ── INDUSTRIAL KPIs ───────────────────────────────────────────────────
    row = _header_block(ws, row, "4. INDUSTRIAL / OPERATIONAL KPIs", "kpi_header", n)

    assum_vol_row = 8
    assum_cap_row = 7
    vol_fmls = [f"=ASSUMPTIONS!{_col(i)}{assum_vol_row}" for i in range(n)]
    cap_fmls = [f"=ASSUMPTIONS!{_col(i)}{assum_cap_row}" for i in range(n)]
    util_fmls = [f"=IFERROR(ASSUMPTIONS!{_col(i)}{assum_vol_row}/ASSUMPTIONS!{_col(i)}{assum_cap_row},\"-\")"
                 for i in range(n)]
    price_fmls = [f"=ASSUMPTIONS!{_col(i)}13" for i in range(n)]

    _data_row(ws, row, "Sales volume", vol_fmls, FONTS["link"], unit="MT", num_fmt="k0")
    row += 1
    _data_row(ws, row, "Installed capacity", cap_fmls, FONTS["link"], unit="MT", num_fmt="k0")
    row += 1
    _data_row(ws, row, "Capacity utilisation", util_fmls, FONTS["formula"], unit="%", num_fmt="pct1")
    row += 1
    _data_row(ws, row, "Average sales price", price_fmls, FONTS["link"], unit="$/MT", num_fmt="k1")
    row += 1

    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# CHECKS
# ═══════════════════════════════════════════════════════════════════════════════
def build_checks(wb, config, years, row_map):
    ws = _sheet_setup(wb, "CHECKS", "A52A2A")
    n = len(years)
    _col_widths(ws, n)
    row = 1

    ws.row_dimensions[row].height = 34
    end = get_column_letter(FIRST_DATA_COL + n - 1)
    merge_and_style(ws, f"B{row}:{end}{row}", "MODEL INTEGRITY CHECKS",
                    FONTS["title"], fill(C["navy"]), AL["center"])
    row += 2
    row = _period_headers(ws, row, years)

    checks = [
        ("Balance sheet balance",
         f"=IF(ABS('BALANCE SHEET'!{{c}}{row_map.get('bs_total_assets',30)}-"
         f"'BALANCE SHEET'!{{c}}{row_map.get('bs_total_l_and_e',50)})<1,\"✓\",\"✗\")"),
        ("P&L net income = BS retained earnings",
         f"=IF(ABS('P&L'!{{c}}{row_map.get('pl_net_income',80)}-"
         f"'BALANCE SHEET'!{{c}}{row_map.get('bs_total_equity',45)})<1,\"~OK\",\"Review\")"),
        ("Closing cash matches BS cash",
         f"=IF(ABS('CASH FLOW'!{{c}}{row_map.get('cf_closing_cash',50)}-"
         f"'BALANCE SHEET'!{{c}}{row_map.get('bs_total_assets',30)})<1,\"✓\",\"Review\")"),
        ("EBITDA >= 0 (warning)",
         f"=IF('P&L'!{{c}}{row_map.get('pl_ebitda_adj',50)}>=0,\"✓\",\"Warning\")"),
        ("FCF >= 0 (warning)",
         f"=IF('CASH FLOW'!{{c}}{row_map.get('cf_fcf',45)}>=0,\"✓\",\"Warning\")"),
        ("Leverage covenant",
         f"=IFERROR('DEBT SCHEDULE'!{'{c}'}{ row_map.get('ds_lev_limit',30)},\"N/A\")"),
    ]

    row = _header_block(ws, row, "CRITICAL CHECKS (✗ = model error)", "kpi_header", n)
    for label, fml_tpl in checks:
        fmls = [fml_tpl.replace("{c}", _col(i)) for i in range(n)]
        ws.row_dimensions[row].height = 17
        c_lbl = ws.cell(row=row, column=LABEL_COL, value=label)
        style_cell(c_lbl, font=FONTS["label"], align=AL["left"])
        for i, f in enumerate(fmls):
            c = ws.cell(row=row, column=FIRST_DATA_COL + i, value=f)
            style_cell(c, font=FONTS["formula"], align=AL["center"])
        row += 1

    return ws
