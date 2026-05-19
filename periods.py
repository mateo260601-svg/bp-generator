from openpyxl.utils import get_column_letter
from styles import *

def build_config(wb, config):
    ws = wb.create_sheet("CONFIG", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    # Column widths
    set_col_width(ws, "A", 3)
    set_col_width(ws, "B", 32)
    set_col_width(ws, "C", 22)
    set_col_width(ws, "D", 22)
    set_col_width(ws, "E", 18)
    set_col_width(ws, "F", 18)
    set_col_width(ws, "G", 18)
    set_col_width(ws, "H", 3)

    row = 1

    # ── Title banner ────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 40
    merge_and_style(ws, f"B{row}:G{row}",
                    "BP GENERATOR — MASTER CONFIGURATION",
                    FONTS["title"], fill(C["navy"]), AL["center"])
    row += 1

    ws.row_dimensions[row].height = 18
    merge_and_style(ws, f"B{row}:G{row}",
                    "All model assumptions flow from this sheet — do not edit other sheets directly",
                    FONTS["note"], fill(C["mid_blue"]), AL["center"])
    row += 2

    # ── SECTION 1: General ──────────────────────────────────────────────────
    ws.row_dimensions[row].height = 20
    merge_and_style(ws, f"B{row}:G{row}", "1. GENERAL PARAMETERS",
                    FONTS["section"], fill(C["assump_hdr"]), AL["left"])
    row += 1

    general_params = [
        ("Company / project name",     config["company_name"],    "text",  ""),
        ("Business type",              config["business_type"],   "text",  "Industriel / LBO / SaaS / Immobilier / Restructuring"),
        ("Reporting currency",         config["currency"],        "text",  "USD / EUR / GBP / AED / BHD"),
        ("Units",                      config["units"],           "text",  "k = thousands  |  m = millions"),
        ("Fiscal year start (month)",  config["fy_start_month"],  "int0",  "e.g. 1 = January"),
        ("Projection start year",      config["proj_start_year"], "int0",  ""),
        ("Number of projection years", config["n_years"],         "int0",  "3 / 5 / 7 / 10"),
        ("Actuals through (year)",     config["actuals_end_year"],"int0",  "Last year of historical data"),
        ("Actuals through (month)",    config["actuals_end_month"],"int0", "1–12"),
        ("Monthly or annual model",    config["freq"],            "text",  "Monthly / Annual"),
    ]

    for label, value, fmt, note in general_params:
        ws.row_dimensions[row].height = 17
        c_label = ws.cell(row=row, column=2, value=label)
        style_cell(c_label, font=FONTS["label"], align=AL["left"])

        c_val = ws.cell(row=row, column=3, value=value)
        style_cell(c_val, font=FONTS["input"], fill_=fill(C["input_bg"]),
                   align=AL["left"], num_fmt=NF[fmt])

        if note:
            c_note = ws.cell(row=row, column=4, value=note)
            style_cell(c_note, font=FONTS["note"], align=AL["left"])

        row += 1

    row += 1

    # ── SECTION 2: Modules ──────────────────────────────────────────────────
    ws.row_dimensions[row].height = 20
    merge_and_style(ws, f"B{row}:G{row}", "2. ACTIVE MODULES (1 = ON / 0 = OFF)",
                    FONTS["section"], fill(C["assump_hdr"]), AL["left"])
    row += 1

    # Headers
    for col_idx, hdr in enumerate(["Module", "Active", "Description"], start=2):
        c = ws.cell(row=row, column=col_idx, value=hdr)
        style_cell(c, font=FONTS["header"], fill_=fill(C["kpi_header"]),
                   align=AL["center"])
    row += 1

    modules = [
        ("P&L",            1, "Income statement — always active"),
        ("Balance Sheet",  1, "Full balance sheet with plug"),
        ("Cash Flow",      1, "Direct / indirect cash flow statement"),
        ("NWC",            1, "Working capital: DSO, DIO, DPO"),
        ("CAPEX",          1, "Investment plan and depreciation schedule"),
        ("Debt Schedule",  config["modules"].get("debt", 1),   "Multi-tranche debt amortisation + interest"),
        ("Tax",            config["modules"].get("tax", 1),    "Corporation tax, deferred tax, loss carry-forward"),
        ("Scenarios",      config["modules"].get("scenarios",1),"Low / Base / High on key assumptions"),
        ("Returns / LBO",  config["modules"].get("returns",0), "IRR, MOIC, equity waterfall (LBO only)"),
        ("Valuation",      config["modules"].get("valuation",0),"DCF + EV/EBITDA exit bridge"),
        ("Consolidation",  config["modules"].get("consol",0),  "Multi-entity consolidation with eliminations"),
        ("Dashboard",      1, "Print-ready KPI summary — auto-populated"),
        ("Output (Print)", 1, "A4 print-ready consolidated output"),
    ]

    for mod_name, active, desc in modules:
        ws.row_dimensions[row].height = 17
        c_mod = ws.cell(row=row, column=2, value=mod_name)
        style_cell(c_mod, font=FONTS["label_bold"], align=AL["left"])

        c_act = ws.cell(row=row, column=3, value=active)
        bg = C["check_ok"] if active else C["light_grey"]
        style_cell(c_act, font=FONTS["input"], fill_=fill(bg),
                   align=AL["center"], num_fmt=NF["int0"])

        c_desc = ws.cell(row=row, column=4, value=desc)
        style_cell(c_desc, font=FONTS["italic"], align=AL["left"])
        row += 1

    row += 1

    # ── SECTION 3: Debt instruments ─────────────────────────────────────────
    ws.row_dimensions[row].height = 20
    merge_and_style(ws, f"B{row}:G{row}", "3. DEBT INSTRUMENTS — SELECT & SIZE",
                    FONTS["section"], fill(C["debt_header"]), AL["left"])
    row += 1

    debt_headers = ["Instrument", "Active", "Amount", "Currency", "Notes"]
    for col_idx, hdr in enumerate(debt_headers, start=2):
        c = ws.cell(row=row, column=col_idx, value=hdr)
        style_cell(c, font=FONTS["header"], fill_=fill(C["kpi_header"]),
                   align=AL["center"])
    row += 1

    debt_groups = {
        "Senior secured": [
            ("Term Loan A — amortissable",       "tla",  1,   150000, "USD", "Quarterly amortisation"),
            ("Term Loan B — bullet",             "tlb",  0,   0,      "USD", "Bullet at maturity"),
            ("Super Senior (SSFA)",              "ss",   0,   0,      "USD", "Priority over TLA/TLB"),
            ("Revolving Credit Facility (RCF)",  "rcf",  1,   40000,  "USD", "Commitment fee on undrawn"),
            ("Capex facility",                   "cxf",  0,   0,      "USD", "Drawn as capex incurred"),
            ("Tranche A1 / A2 split-entity",     "a1a2", 0,   0,      "USD", "Activate for multi-entity"),
            ("Islamic — Murabaha",               "mur",  1,   60000,  "USD", "Cost-plus Islamic financing"),
            ("Islamic — Ijara / Wakala",         "ija",  0,   0,      "USD", "Lease-based Islamic financing"),
            ("Multi-currency facility",          "mcf",  0,   0,      "USD", "FX rate in Assumptions"),
        ],
        "Mezzanine / Subordinated": [
            ("Mezzanine — cash pay",             "mez",  0,   0,      "USD", ""),
            ("PIK toggle",                       "pik",  0,   0,      "USD", "Accrues if cash unavailable"),
            ("Unitranche",                       "uni",  0,   0,      "USD", "Blended TLA + mezz rate"),
            ("Second lien",                      "sl",   0,   0,      "USD", ""),
            ("Shareholder loan (SHL)",           "shl",  0,   0,      "USD", "Typically full PIK"),
            ("Vendor loan / seller note",        "vnd",  0,   0,      "USD", ""),
            ("Convertible loan note (CLN)",      "cln",  0,   0,      "USD", "Equity-linked"),
            ("Preferred equity / redeemable",    "pref", 0,   0,      "USD", ""),
            ("Earn-out structuré",               "ern",  0,   0,      "USD", "Contingent on performance"),
            ("Islamic — Mudaraba / Musharaka",   "mud",  0,   0,      "USD", "Profit-sharing Islamic"),
        ],
        "Capital markets": [
            ("High Yield Bond",                  "hyb",  0,   0,      "USD", "Step-up coupon option"),
            ("Senior Secured Notes",             "ssn",  0,   0,      "USD", ""),
            ("Subordinated Notes",               "subn", 0,   0,      "USD", ""),
            ("Convertible bond",                 "cvb",  0,   0,      "USD", ""),
            ("Euro PP / Schuldschein",           "epp",  0,   0,      "EUR", ""),
            ("USPP (US Private Placement)",      "uspp", 0,   0,      "USD", ""),
            ("Sukuk",                            "suk",  0,   0,      "USD", "Islamic capital markets"),
            ("Green / sustainability bond",      "gsb",  0,   0,      "USD", ""),
        ],
        "Specialised / Project": [
            ("Project finance — non-recourse",   "pf",   0,   0,      "USD", "DSCR covenant tracking"),
            ("Bridge loan",                      "bri",  0,   0,      "USD", "Short-term, refinanced"),
            ("Construction facility",            "con",  0,   0,      "USD", "Drawn during build period"),
            ("Leasing financier / opérationnel", "lea",  0,   0,      "USD", ""),
            ("Crédit-bail immobilier",           "cbi",  0,   0,      "EUR", ""),
            ("ECA / State-backed financing",     "eca",  0,   0,      "USD", ""),
            ("Amend & extend (A&E)",             "ae",   0,   0,      "USD", "Maturity extension"),
        ],
        "Restructuring": [
            ("New money facility (DIP-style)",   "dip",  0,   0,      "USD", "Super priority"),
            ("Debt-to-equity swap",              "d2e",  0,   0,      "USD", "Modelled as haircut + equity"),
            ("PIK forcé sur période",            "pikf", 0,   0,      "USD", "Override cash pay"),
            ("Moratorium / standstill",          "mor",  0,   0,      "USD", "Interest freeze period"),
            ("Covenant waiver / reset",          "cwv",  0,   0,      "USD", "New covenant levels"),
        ],
    }

    debt_config = config.get("debt", {})

    for group_name, instruments in debt_groups.items():
        ws.row_dimensions[row].height = 17
        merge_and_style(ws, f"B{row}:G{row}", group_name,
                        FONTS["subheader"], fill(C["debt_sub"]), AL["left"])
        row += 1

        for label, key, default_active, default_amt, default_ccy, notes in instruments:
            ws.row_dimensions[row].height = 16
            active_val = debt_config.get(key, {}).get("active", default_active)
            amt_val    = debt_config.get(key, {}).get("amount", default_amt)
            ccy_val    = debt_config.get(key, {}).get("currency", default_ccy)

            c1 = ws.cell(row=row, column=2, value=label)
            style_cell(c1, font=FONTS["label"], align=AL["left"])

            c2 = ws.cell(row=row, column=3, value=active_val)
            bg = C["check_ok"] if active_val else C["light_grey"]
            style_cell(c2, font=FONTS["input"], fill_=fill(bg),
                       align=AL["center"], num_fmt=NF["int0"])

            c3 = ws.cell(row=row, column=4, value=amt_val if active_val else None)
            style_cell(c3, font=FONTS["input"], fill_=fill(C["input_bg"]) if active_val else fill(C["light_grey"]),
                       align=AL["right"], num_fmt=NF["k0"])

            c4 = ws.cell(row=row, column=5, value=ccy_val)
            style_cell(c4, font=FONTS["label"], align=AL["center"])

            c5 = ws.cell(row=row, column=6, value=notes)
            style_cell(c5, font=FONTS["note"], align=AL["left"])

            row += 1

    row += 1

    # ── SECTION 4: Debt mechanics ───────────────────────────────────────────
    ws.row_dimensions[row].height = 20
    merge_and_style(ws, f"B{row}:G{row}", "4. DEBT MECHANICS & CROSS-CUTTING FEATURES",
                    FONTS["section"], fill(C["debt_header"]), AL["left"])
    row += 1

    mechanics = [
        ("Base rate index",             config.get("base_rate_index","SOFR"),      "text",  "EURIBOR 3M / EURIBOR 6M / SOFR / SONIA / Fixed"),
        ("Base rate (%)",               config.get("base_rate_pct", 4.50),         "pct2",  "Current market rate — update per period in Assumptions"),
        ("Base rate floor (%)",         config.get("base_rate_floor", 0.0),        "pct2",  "Typically 0% — SOFR/EURIBOR floor"),
        ("TLA margin (%)",              config.get("tla_margin", 3.25),            "pct2",  ""),
        ("TLB margin (%)",              config.get("tlb_margin", 4.00),            "pct2",  ""),
        ("Mezz cash margin (%)",        config.get("mezz_cash_margin", 7.00),      "pct2",  ""),
        ("Mezz PIK margin (%)",         config.get("mezz_pik_margin", 4.00),       "pct2",  ""),
        ("Upfront / arrangement fee (%)",config.get("upfront_fee_pct", 1.50),      "pct2",  "Amortised over tenor"),
        ("Commitment fee on undrawn (%)",config.get("commit_fee_pct", 0.50),       "pct2",  "Applied to undrawn RCF"),
        ("OID (%)",                     config.get("oid_pct", 0.0),               "pct2",  "Original issue discount — 0 if none"),
        ("Cash sweep (%)",              config.get("cash_sweep_pct", 75),          "pct1",  "% of excess cash applied to mandatory prepayment"),
        ("Cash sweep active (1/0)",     config.get("cash_sweep_active", 1),        "int0",  ""),
        ("Margin ratchet active (1/0)", config.get("margin_ratchet", 0),           "int0",  "Margin steps down as leverage improves"),
        ("Covenant tracking (1/0)",     config.get("covenant_tracking", 1),        "int0",  ""),
        ("Leverage covenant (max x)",   config.get("lev_covenant", 5.0),           "mult",  "Net debt / EBITDA"),
        ("Interest cover covenant (min x)",config.get("icr_covenant", 2.0),        "mult",  "EBITDA / Net interest"),
        ("DSCR covenant (min x)",       config.get("dscr_covenant", 1.0),          "mult",  "For project finance — set 0 if N/A"),
        ("FX hedge active (1/0)",       config.get("fx_hedge", 0),                "int0",  ""),
        ("Interest rate swap (1/0)",    config.get("irs_active", 0),              "int0",  "Fixed/floating swap"),
    ]

    for label, value, fmt, note in mechanics:
        ws.row_dimensions[row].height = 17
        c_label = ws.cell(row=row, column=2, value=label)
        style_cell(c_label, font=FONTS["label"], align=AL["left"])

        c_val = ws.cell(row=row, column=3, value=value)
        style_cell(c_val, font=FONTS["input"], fill_=fill(C["input_bg"]),
                   align=AL["right"], num_fmt=NF[fmt])

        if note:
            c_note = ws.cell(row=row, column=4, value=note)
            style_cell(c_note, font=FONTS["note"], align=AL["left"])
        row += 1

    row += 1

    # ── SECTION 5: Colour coding legend ─────────────────────────────────────
    ws.row_dimensions[row].height = 20
    merge_and_style(ws, f"B{row}:G{row}", "5. COLOUR CODING — READ BEFORE EDITING",
                    FONTS["section"], fill(C["assump_hdr"]), AL["left"])
    row += 1

    legend = [
        (C["input_bg"],    C["input_blue"],  "Blue text on yellow  →  hardcoded input — edit freely"),
        (C["white"],       C["formula_clr"], "Black text on white  →  formula — do NOT override"),
        (C["white"],       C["link_clr"],    "Green text on white  →  cross-sheet link — do NOT override"),
        (C["check_ok"],    C["black"],       "Green fill  →  check OK / module active"),
        (C["check_err"],   C["black"],       "Red fill    →  check error / balance mismatch"),
        (C["total_fill"],  C["white"],       "Dark blue fill  →  major total (EBITDA, Net Income, Net Debt)"),
        (C["subtot_fill"], C["navy"],        "Light blue fill  →  sub-total"),
    ]

    for bg, txt, desc in legend:
        ws.row_dimensions[row].height = 17
        c_swatch = ws.cell(row=row, column=2, value="   ")
        style_cell(c_swatch, fill_=fill(bg), align=AL["center"])

        c_desc = ws.cell(row=row, column=3, value=desc)
        style_cell(c_desc, font=Font(name="Calibri", size=9, color=txt, bold=False),
                   fill_=fill(bg), align=AL["left"])

        merge_and_style(ws, f"C{row}:G{row}", desc,
                        Font(name="Calibri", size=9, color=txt),
                        fill(bg), AL["left"])
        row += 1

    # Tab colour
    ws.sheet_properties.tabColor = "0D1B3E"
    return ws
