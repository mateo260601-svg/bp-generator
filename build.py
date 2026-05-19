from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Palette ────────────────────────────────────────────────────────────────────
C = {
    # Backgrounds
    "navy":        "0D1B3E",
    "dark_blue":   "1A3560",
    "mid_blue":    "2E5F9E",
    "light_blue":  "C5D8F0",
    "pale_blue":   "EBF2FA",
    "white":       "FFFFFF",
    "off_white":   "F7F9FC",
    "light_grey":  "F0F0F0",
    "mid_grey":    "D0D0D0",
    "dark_grey":   "606060",
    "black":       "000000",
    # Section headers by statement
    "pl_header":   "1A3560",   # P&L  → dark blue
    "bs_header":   "1E4D2B",   # BS   → dark green
    "cf_header":   "4A235A",   # CF   → dark purple
    "nwc_header":  "7B3F00",   # NWC  → dark brown
    "debt_header": "7B0000",   # Debt → dark red
    "kpi_header":  "2C4A6E",   # KPIs → slate
    "assump_hdr":  "0D3349",   # Assumptions
    # Sub-section fills
    "pl_sub":      "D6E4F0",
    "bs_sub":      "D6EAD6",
    "cf_sub":      "E8D5F0",
    "nwc_sub":     "FAE5D3",
    "debt_sub":    "FAD5D5",
    "kpi_sub":     "D5E3F0",
    "assump_sub":  "D0E8F5",
    # Total rows
    "total_fill":  "2E5F9E",
    "subtot_fill": "C5D8F0",
    "ebitda_fill": "1A3560",
    "check_ok":    "C6EFCE",
    "check_err":   "FFC7CE",
    # Input cells
    "input_bg":    "FFF2CC",   # yellow
    "input_blue":  "0000FF",   # blue text = hardcoded input
    "formula_clr": "000000",   # black = formula
    "link_clr":    "008000",   # green = cross-sheet link
}

# ── Fonts ──────────────────────────────────────────────────────────────────────
def fnt(bold=False, size=9, color="000000", italic=False, name="Calibri"):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

FONTS = {
    "title":     fnt(bold=True,  size=14, color=C["white"]),
    "subtitle":  fnt(bold=True,  size=10, color=C["white"]),
    "header":    fnt(bold=True,  size=9,  color=C["white"]),
    "subheader": fnt(bold=True,  size=9,  color=C["navy"]),
    "section":   fnt(bold=True,  size=9,  color=C["white"]),
    "label":     fnt(bold=False, size=9,  color=C["black"]),
    "label_bold":fnt(bold=True,  size=9,  color=C["black"]),
    "input":     fnt(bold=False, size=9,  color=C["input_blue"]),
    "formula":   fnt(bold=False, size=9,  color=C["formula_clr"]),
    "link":      fnt(bold=False, size=9,  color=C["link_clr"]),
    "total":     fnt(bold=True,  size=9,  color=C["white"]),
    "subtotal":  fnt(bold=True,  size=9,  color=C["navy"]),
    "italic":    fnt(bold=False, size=9,  color=C["dark_grey"], italic=True),
    "note":      fnt(bold=False, size=8,  color=C["dark_grey"], italic=True),
}

# ── Fills ──────────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

FILLS = {k: fill(v) for k, v in C.items()}

# ── Borders ────────────────────────────────────────────────────────────────────
def _side(style=None):
    return Side(style=style, color="000000") if style else Side(style=None)

BORDERS = {
    "none":   Border(),
    "thin":   Border(bottom=_side("thin")),
    "thick":  Border(bottom=_side("medium")),
    "box":    Border(left=_side("thin"), right=_side("thin"),
                     top=_side("thin"), bottom=_side("thin")),
    "top":    Border(top=_side("thin")),
    "bottom": Border(bottom=_side("thin")),
    "double_bot": Border(bottom=_side("double")),
}

# ── Alignments ─────────────────────────────────────────────────────────────────
AL = {
    "left":    Alignment(horizontal="left",   vertical="center", wrap_text=False),
    "center":  Alignment(horizontal="center", vertical="center", wrap_text=False),
    "right":   Alignment(horizontal="right",  vertical="center", wrap_text=False),
    "left_w":  Alignment(horizontal="left",   vertical="center", wrap_text=True),
    "center_w":Alignment(horizontal="center", vertical="center", wrap_text=True),
}

# ── Number formats ─────────────────────────────────────────────────────────────
NF = {
    "k0":    '#,##0;(#,##0);"-"',           # thousands, no decimal
    "k1":    '#,##0.0;(#,##0.0);"-"',       # thousands, 1 dp
    "m0":    '#,##0;(#,##0);"-"',           # millions  (same mask, units in header)
    "pct1":  '0.0%;(0.0%);"-"',             # percentage 1 dp
    "pct2":  '0.00%;(0.00%);"-"',           # percentage 2 dp
    "mult":  '0.0"x";(0.0"x");"-"',         # multiples
    "date":  'MMM-YY',                       # month header
    "year":  '"FY"YYYY',                     # year header
    "int0":  '#,##0;(#,##0);"-"',           # integer
    "text":  '@',                            # text
    "check": '"✓";;"✗"',                   # boolean check
}

# ── Helpers ────────────────────────────────────────────────────────────────────
def style_cell(cell, font=None, fill_=None, border=None, align=None, num_fmt=None):
    if font:    cell.font      = font
    if fill_:   cell.fill      = fill_
    if border:  cell.border    = border
    if align:   cell.alignment = align
    if num_fmt: cell.number_format = num_fmt

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def merge_and_style(ws, cell_range, value, font, fill_, align=None, border=None):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = value
    c.font  = font
    c.fill  = fill_
    if align:  c.alignment = align
    if border: c.border    = border

def row_height(ws, row, height):
    ws.row_dimensions[row].height = height
