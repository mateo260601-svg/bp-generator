"""
Scenarios Engine
Generates Low Case / Base Case / Best Case assumptions
and injects them as dedicated tabs in the BP Excel output.
"""
from openpyxl.utils import get_column_letter
from styles import *

# ── Default scenario deltas ──────────────────────────────────────────────────
# Each scenario applies multipliers/deltas vs Base Case
SCENARIO_DEFS = {
    "low": {
        "label":       "Low Case",
        "tab_color":   "C0392B",   # red
        "fill_header": "7B0000",
        "fill_sub":    "FAD5D5",
        "revenue_delta":      -0.15,   # -15% vs base
        "ebitda_margin_delta": -0.05,  # -5pp margin compression
        "volume_delta":       -0.12,
        "price_delta":        -0.08,
        "cost_inflation":     +0.03,   # +3% cost creep
        "capex_delta":        +0.10,   # +10% capex (contingency)
        "dso_delta":          +10,     # +10 days (slower collections)
        "dpo_delta":          -5,      # -5 days (faster payments)
        "dio_delta":          +8,      # +8 days (inventory build)
        "description": "Environnement dégradé — tensions macro, pression prix, coûts élevés",
    },
    "base": {
        "label":       "Base Case",
        "tab_color":   "2E5F9E",   # navy blue
        "fill_header": "1A3560",
        "fill_sub":    "C5D8F0",
        "revenue_delta":       0.0,
        "ebitda_margin_delta": 0.0,
        "volume_delta":        0.0,
        "price_delta":         0.0,
        "cost_inflation":      0.0,
        "capex_delta":         0.0,
        "dso_delta":           0,
        "dpo_delta":           0,
        "dio_delta":           0,
        "description": "Hypothèses centrales — trajectoire de marché normalisée",
    },
    "best": {
        "label":       "Best Case",
        "tab_color":   "27AE60",   # green
        "fill_header": "1E4D2B",
        "fill_sub":    "D6EAD6",
        "revenue_delta":      +0.15,
        "ebitda_margin_delta": +0.04,
        "volume_delta":       +0.12,
        "price_delta":        +0.08,
        "cost_inflation":     -0.02,
        "capex_delta":        -0.05,
        "dso_delta":          -8,
        "dpo_delta":          +5,
        "dio_delta":          -5,
        "description": "Scénario favorable — accélération volumes, pricing power, optimisation coûts",
    },
}

def _col(first_data_col, idx):
    return get_column_letter(first_data_col + idx)

def _apply_delta(base_list, delta_pct):
    """Apply a percentage delta to a list of values."""
    return [round(v * (1 + delta_pct)) if v else v for v in base_list]

def _apply_additive(base_list, delta):
    """Apply an additive delta (e.g. days) to a list."""
    return [max(0, v + delta) if v is not None else v for v in base_list]

def build_scenario_assumptions(base_config: dict, scenario_key: str) -> dict:
    """
    Takes base config and returns a modified config for the given scenario.
    """
    import copy
    cfg = copy.deepcopy(base_config)
    s = SCENARIO_DEFS[scenario_key]
    n = cfg["n_years"]

    rev  = cfg.get("revenue", {})
    cost = cfg.get("costs",   {})
    nwc  = cfg.get("nwc",     {})
    cpx  = cfg.get("capex",   {})

    rd = s["revenue_delta"]
    vd = s["volume_delta"]
    pd = s["price_delta"]
    ci = s["cost_inflation"]
    cd = s["capex_delta"]

    # Revenue drivers
    if "base_volume" in rev:
        rev["base_volume"]  = _apply_delta(rev["base_volume"],  vd)
    if "price_per_mt" in rev:
        rev["price_per_mt"] = _apply_delta(rev["price_per_mt"], pd)
    if "capacity_mt" in rev:
        pass  # capacity unchanged

    # Cost drivers — inflate/deflate variable costs
    for cost_key in ["direct_mat_mt", "utilities_mt", "packing_mt", "var_opex_other_mt"]:
        if cost_key in cost:
            cost[cost_key] = _apply_delta(cost[cost_key], ci)

    # Fixed costs — partial sensitivity
    for fixed_key in ["staff_prod", "staff_sga", "maintenance"]:
        if fixed_key in cost:
            cost[fixed_key] = _apply_delta(cost[fixed_key], ci * 0.5)

    # NWC days
    for k, delta in [("dso", s["dso_delta"]), ("dpo", s["dpo_delta"]), ("dio", s["dio_delta"])]:
        if k in nwc:
            nwc[k] = _apply_additive(nwc[k], delta)

    # Capex
    for k in ["maint", "expan"]:
        if k in cpx:
            cpx[k] = _apply_delta(cpx[k], cd)

    cfg["revenue"]  = rev
    cfg["costs"]    = cost
    cfg["nwc"]      = nwc
    cfg["capex"]    = cpx
    cfg["_scenario"] = scenario_key
    cfg["_scenario_label"] = s["label"]

    return cfg

def build_scenarios_sheet(wb, base_config: dict, years: list,
                           base_row_map: dict) -> None:
    """
    Builds a SCENARIOS sheet with Low / Base / Best side-by-side comparison.
    Called after all other sheets are built.
    """
    from periods import build_periods

    ws = wb.create_sheet("SCENARIOS")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.tabColor = "2C4A6E"

    n = len(years)
    FIRST = 5
    LABEL_COL = 2

    # Column widths
    set_col_width(ws, "A", 2)
    set_col_width(ws, "B", 34)
    set_col_width(ws, "C", 10)
    for i in range(n):
        set_col_width(ws, get_column_letter(FIRST + i), 13)
    set_col_width(ws, get_column_letter(FIRST + n), 2)

    row = 1

    # ── Title ────────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 38
    end = get_column_letter(FIRST + n - 1)
    merge_and_style(ws, f"B{row}:{end}{row}",
                    f"{base_config['company_name']} — Analyse de Scénarios",
                    FONTS["title"], fill(C["navy"]), AL["center"])
    row += 1

    ws.row_dimensions[row].height = 16
    merge_and_style(ws, f"B{row}:{end}{row}",
                    "Low Case  ·  Base Case  ·  Best Case — sensibilité des hypothèses clés",
                    FONTS["note"], fill(C["mid_blue"]), AL["center"])
    row += 2

    # ── For each scenario: compute derived values ────────────────────────────
    from build import build_workbook
    import tempfile, os

    scenario_data = {}
    for sk in ["low", "base", "best"]:
        sc_cfg = build_scenario_assumptions(base_config, sk)
        rev    = sc_cfg.get("revenue", {})
        cost   = sc_cfg.get("costs",   {})
        nwc    = sc_cfg.get("nwc",     {})
        cpx    = sc_cfg.get("capex",   {})
        n_yrs  = sc_cfg["n_years"]

        vols   = rev.get("base_volume",  [0]*n_yrs)
        prices = rev.get("price_per_mt", [0]*n_yrs)
        frt    = rev.get("freight_mt",   [35]*n_yrs)
        comm   = rev.get("commission_pct",[0.01]*n_yrs)

        gross_rev = [v*p/1000 for v,p in zip(vols, prices)]
        net_rev   = [g - v*f/1000 - g*c for g,v,f,c in zip(gross_rev,vols,frt,comm)]

        dm    = [v*cost.get("direct_mat_mt",[520]*n_yrs)[i]/1000 for i,v in enumerate(vols)]
        util  = [v*cost.get("utilities_mt", [45]*n_yrs)[i]/1000  for i,v in enumerate(vols)]
        pack  = [v*cost.get("packing_mt",   [18]*n_yrs)[i]/1000  for i,v in enumerate(vols)]
        var_o = [v*cost.get("var_opex_other_mt",[12]*n_yrs)[i]/1000 for i,v in enumerate(vols)]
        cogs  = [dm[i]+util[i]+pack[i]+var_o[i] for i in range(n_yrs)]
        gp    = [net_rev[i]-cogs[i] for i in range(n_yrs)]

        sp    = cost.get("staff_prod",[8500]*n_yrs)
        ss    = cost.get("staff_sga", [3200]*n_yrs)
        maint = cost.get("maintenance",[2400]*n_yrs)
        ins   = cost.get("insurance", [800]*n_yrs)
        rent  = cost.get("rent",      [600]*n_yrs)
        fixed = [sp[i]+ss[i]+maint[i]+ins[i]+rent[i] for i in range(n_yrs)]

        ebitda = [gp[i]-fixed[i] for i in range(n_yrs)]
        ebitda_mgn = [e/r if r else 0 for e,r in zip(ebitda, net_rev)]

        dep_open = sc_cfg.get("capex",{}).get("opening_ppe",95000)
        dep_life = sc_cfg.get("capex",{}).get("useful_life",20)
        dep_ann  = [dep_open/dep_life]*n_yrs
        ebit     = [ebitda[i]-dep_ann[i] for i in range(n_yrs)]

        capex_t  = [cpx.get("maint",[3500]*n_yrs)[i]+cpx.get("expan",[5000]*n_yrs)[i]
                    for i in range(n_yrs)]

        scenario_data[sk] = {
            "net_rev":    net_rev,
            "gp":         gp,
            "ebitda":     ebitda,
            "ebitda_mgn": ebitda_mgn,
            "ebit":       ebit,
            "capex":      capex_t,
            "dso":        nwc.get("dso",[45]*n_yrs),
            "dpo":        nwc.get("dpo",[55]*n_yrs),
        }

    # ── Scenario descriptions ─────────────────────────────────────────────────
    for sk, sdef in SCENARIO_DEFS.items():
        ws.row_dimensions[row].height = 22
        fh = sdef["fill_header"]
        merge_and_style(ws, f"B{row}:C{row}", sdef["label"],
                        FONTS["section"], fill(fh), AL["left"])
        merge_and_style(ws, f"D{row}:{end}{row}", sdef["description"],
                        FONTS["note"], fill(fh), AL["left"])
        row += 1
    row += 1

    # ── Scenario comparison table ─────────────────────────────────────────────
    # For each KPI, show all 3 scenarios stacked
    ccy  = base_config.get("currency","USD")
    unit = f"{ccy}k"

    KPI_ROWS = [
        ("Revenus nets",       "net_rev",    "k0",   unit),
        ("Marge brute",        "gp",         "k0",   unit),
        ("EBITDA",             "ebitda",     "k0",   unit),
        ("Marge EBITDA %",     "ebitda_mgn", "pct1", "%"),
        ("EBIT",               "ebit",       "k0",   unit),
        ("Capex total",        "capex",      "k0",   unit),
        ("DSO",                "dso",        "int0", "jours"),
        ("DPO",                "dpo",        "int0", "jours"),
    ]

    for kpi_label, kpi_key, fmt_key, kpi_unit in KPI_ROWS:
        # KPI header
        ws.row_dimensions[row].height = 18
        merge_and_style(ws, f"B{row}:{end}{row}", kpi_label,
                        FONTS["subheader"], fill(C["pale_blue"]), AL["left"])
        row += 1

        # Year headers
        ws.row_dimensions[row].height = 15
        c_h = ws.cell(row=row, column=LABEL_COL, value="Scénario")
        style_cell(c_h, font=FONTS["header"], fill_=fill(C["kpi_header"]),
                   align=AL["left"])
        c_u = ws.cell(row=row, column=3, value=kpi_unit)
        style_cell(c_u, font=FONTS["header"], fill_=fill(C["kpi_header"]),
                   align=AL["center"])
        for i, yr in enumerate(years):
            c = ws.cell(row=row, column=FIRST+i, value=yr["fy"])
            style_cell(c, font=FONTS["header"], fill_=fill(C["kpi_header"]),
                       align=AL["center"])
        row += 1

        # One row per scenario
        for sk in ["low", "base", "best"]:
            sdef = SCENARIO_DEFS[sk]
            vals = scenario_data[sk].get(kpi_key, [0]*n)
            fh   = sdef["fill_header"]
            fs   = sdef["fill_sub"]
            is_base = sk == "base"

            ws.row_dimensions[row].height = 16

            c_lbl = ws.cell(row=row, column=LABEL_COL,
                            value=sdef["label"])
            style_cell(c_lbl,
                       font=FONTS["total"] if is_base else FONTS["label_bold"],
                       fill_=fill(fh) if is_base else fill(fs),
                       align=AL["left"])

            c_u2 = ws.cell(row=row, column=3, value=kpi_unit)
            style_cell(c_u2,
                       font=FONTS["italic"],
                       fill_=fill(fh) if is_base else fill(fs),
                       align=AL["center"])

            for i, val in enumerate(vals[:n]):
                c = ws.cell(row=row, column=FIRST+i, value=round(val, 4))
                style_cell(c,
                           font=FONTS["total"] if is_base else FONTS["formula"],
                           fill_=fill(fh) if is_base else fill(fs),
                           align=AL["right"],
                           num_fmt=NF[fmt_key])
            row += 1

        # Delta row: Best vs Low
        ws.row_dimensions[row].height = 14
        c_d = ws.cell(row=row, column=LABEL_COL, value="  Écart Best / Low")
        style_cell(c_d, font=FONTS["italic"], align=AL["left"])
        low_vals  = scenario_data["low"].get(kpi_key,  [0]*n)
        best_vals = scenario_data["best"].get(kpi_key, [0]*n)
        for i in range(n):
            delta = best_vals[i] - low_vals[i] if i < len(best_vals) else 0
            c = ws.cell(row=row, column=FIRST+i, value=round(delta, 4))
            style_cell(c, font=FONTS["italic"], align=AL["right"],
                       num_fmt=NF[fmt_key])
        row += 2

    # ── EBITDA bridge chart data (for reference) ──────────────────────────────
    ws.row_dimensions[row].height = 20
    merge_and_style(ws, f"B{row}:{end}{row}",
                    "HYPOTHÈSES PAR SCÉNARIO — COMPARATIF",
                    FONTS["section"], fill(C["navy"]), AL["left"])
    row += 1

    # Table of deltas vs base
    delta_params = [
        ("Revenus vs Base",         "revenue_delta",      "pct1"),
        ("Volume vs Base",          "volume_delta",       "pct1"),
        ("Prix vs Base",            "price_delta",        "pct1"),
        ("Inflation coûts vs Base", "cost_inflation",     "pct1"),
        ("Capex vs Base",           "capex_delta",        "pct1"),
        ("Delta DSO (jours)",       "dso_delta",          "int0"),
        ("Delta DPO (jours)",       "dpo_delta",          "int0"),
        ("Delta DIO (jours)",       "dio_delta",          "int0"),
    ]

    # Headers
    ws.row_dimensions[row].height = 16
    c_h = ws.cell(row=row, column=LABEL_COL, value="Paramètre")
    style_cell(c_h, font=FONTS["header"], fill_=fill(C["kpi_header"]), align=AL["left"])
    for i, sk in enumerate(["low","base","best"]):
        c = ws.cell(row=row, column=FIRST+i, value=SCENARIO_DEFS[sk]["label"])
        style_cell(c, font=FONTS["header"],
                   fill_=fill(SCENARIO_DEFS[sk]["fill_header"]), align=AL["center"])
    row += 1

    for param_label, param_key, fmt_key in delta_params:
        ws.row_dimensions[row].height = 16
        c_lbl = ws.cell(row=row, column=LABEL_COL, value=param_label)
        style_cell(c_lbl, font=FONTS["label"], align=AL["left"])
        for i, sk in enumerate(["low","base","best"]):
            val = SCENARIO_DEFS[sk][param_key]
            c = ws.cell(row=row, column=FIRST+i, value=val)
            fs = SCENARIO_DEFS[sk]["fill_sub"]
            style_cell(c, font=FONTS["formula"],
                       fill_=fill(fs), align=AL["right"],
                       num_fmt=NF[fmt_key])
        row += 1

    return ws
