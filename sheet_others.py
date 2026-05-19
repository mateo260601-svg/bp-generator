"""
Preview Engine v2
Reads numeric values directly from ASSUMPTIONS (hardcoded inputs),
computes full P&L / KPIs / debt for inline browser display.
No formula evaluation needed.
"""
import re
import openpyxl


def _label(row):
    return str(row[1] or row[0] or "").strip()


def _nums(row, start=3, n=12):
    out = []
    for ci in range(start, min(len(row), start + n)):
        v = row[ci]
        out.append(round(float(v), 4) if isinstance(v, (int, float)) else None)
    return out


def _read_sheet(ws, label_col=1, data_start_col=3, max_rows=80):
    data = {}
    for row in ws.iter_rows(max_row=max_rows, values_only=True):
        lbl = str(row[label_col] or "").strip()
        if not lbl:
            continue
        vals = _nums(row, start=data_start_col)
        if any(v is not None for v in vals):
            data[lbl] = vals
    return data


def _g(d, key, n, default=0.0):
    """Get n values from dict key, filling with default."""
    v = d.get(key, [])
    padded = [x if x is not None else default for x in v] + [default] * n
    return padded[:n]


def _div(a, b):
    return round(a / b, 4) if b and b != 0 else None


def build_preview(filepath: str) -> dict:
    wb    = openpyxl.load_workbook(filepath, data_only=True)
    sheets = wb.sheetnames

    if "ASSUMPTIONS" not in sheets:
        return {"error": "ASSUMPTIONS introuvable", "sheets": sheets}

    A = _read_sheet(wb["ASSUMPTIONS"], label_col=1, data_start_col=4, max_rows=70)

    # Detect n from revenue row
    rev_raw = A.get("Gross revenue") or A.get("Net realisation") or []
    n = sum(1 for v in rev_raw if v is not None)
    n = max(n, 3)

    # Year labels
    years = []
    for row in wb["ASSUMPTIONS"].iter_rows(max_row=6, values_only=True):
        for v in row:
            m = re.search(r"FY\d{4}", str(v or ""))
            if m and str(v) not in years:
                years.append(str(v))
    years = years[:n]
    if not years:
        import datetime
        years = [f"FY{datetime.datetime.now().year + i}" for i in range(n)]

    def g(k, d=0.0):
        return _g(A, k, n, d)

    # ── Revenue ───────────────────────────────────────────────────────────────
    vol      = g("Sales volume")
    price_mt = g("Average sales price", 1050)
    frt_mt   = g("Freight & forwarding", 35)
    comm_pct = g("Sales commission", 0.01)
    capacity = g("Installed capacity", 220000)

    gross_rev = A.get("Gross revenue") or [vol[i]*price_mt[i]/1000 for i in range(n)]
    gross_rev = [gross_rev[i] if i < len(gross_rev) and gross_rev[i] is not None
                 else vol[i]*price_mt[i]/1000 for i in range(n)]

    net_rev = A.get("Net realisation") or []
    net_rev = [net_rev[i] if i < len(net_rev) and net_rev[i] is not None
               else gross_rev[i] - vol[i]*frt_mt[i]/1000 - gross_rev[i]*comm_pct[i]
               for i in range(n)]

    # ── COGS ──────────────────────────────────────────────────────────────────
    dm_mt   = g("Direct materials", 520)
    util_mt = g("Utilities", 45)
    pack_mt = g("Packing costs", 18)
    var_mt  = g("Variable OpEx — other", 12)

    dm   = [vol[i]*dm_mt[i]/1000   for i in range(n)]
    util = [vol[i]*util_mt[i]/1000 for i in range(n)]
    pack = [vol[i]*pack_mt[i]/1000 for i in range(n)]
    var_ = [vol[i]*var_mt[i]/1000  for i in range(n)]
    cogs = [dm[i]+util[i]+pack[i]+var_[i] for i in range(n)]
    gp   = [net_rev[i]-cogs[i] for i in range(n)]

    # ── Fixed OpEx ────────────────────────────────────────────────────────────
    sp   = g("Personnel — production", 8500)
    ss   = g("Personnel — SG&A", 3200)
    mnt  = g("Maintenance & repairs", 2400)
    ins  = g("Insurance", 800)
    rent = g("Rent / site costs", 600)
    it_s = g("IT & systems", 400)
    prof = g("Professional fees", 350)
    osga = g("Other SG&A", 500)
    restr= g("Restructuring / one-off", 0)

    fixed = [sp[i]+ss[i]+mnt[i]+ins[i]+rent[i]+it_s[i]+prof[i]+osga[i] for i in range(n)]
    ebitda     = [gp[i]-fixed[i] for i in range(n)]
    adj_ebitda = [ebitda[i]+restr[i] for i in range(n)]

    # ── D&A ───────────────────────────────────────────────────────────────────
    ul  = (A.get("Depreciation — useful life (yrs)") or [20])[0] or 20
    ppe = (A.get("Opening PP&E (gross)") or [95000])[0] or 95000
    dep = [round(ppe / ul)] * n
    ebit = [adj_ebitda[i]-dep[i] for i in range(n)]

    # ── Capex ─────────────────────────────────────────────────────────────────
    maint_c = g("Maintenance capex", 3500)
    expan_c = g("Expansion capex", 5000)
    capex_t = [maint_c[i]+expan_c[i] for i in range(n)]

    # ── Tax & macro ───────────────────────────────────────────────────────────
    tax_r = g("Corporation tax rate", 0.25)
    dso   = g("Days Sales Outstanding (DSO)", 50)
    dpo   = g("Days Payables Outstanding (DPO)", 55)
    dio   = g("Days Inventory Outstanding (DIO)", 35)

    # ── Debt ──────────────────────────────────────────────────────────────────
    int_exp    = [0.0] * n
    gross_debt = [0.0] * n
    repay      = [0.0] * n
    debt_rows  = []

    if "DEBT SCHEDULE" in sheets:
        D = _read_sheet(wb["DEBT SCHEDULE"], label_col=1, data_start_col=3, max_rows=60)
        for lbl, vals in D.items():
            padded = [(v or 0) for v in (vals + [0]*n)][:n]
            if "Total gross debt" in lbl:
                gross_debt = padded
            elif "Total cash interest" in lbl:
                int_exp = padded
            elif "Total repayment" in lbl:
                repay = padded
            elif any(k in lbl for k in ["Term Loan","Senior","RCF","Murabaha","Mezz",
                                         "PIK","Unitranche","Bond","Shareholder"]):
                b0 = padded[0]
                if b0 > 1000:
                    debt_rows.append({"label": lbl, "balances": padded})

        # Estimate interest if not found
        if all(v == 0 for v in int_exp) and any(v > 0 for v in gross_debt):
            int_exp = [round(gross_debt[i] * 0.075) for i in range(n)]

    pbt     = [ebit[i]-int_exp[i] for i in range(n)]
    tax_chg = [max(0, round(pbt[i]*tax_r[i])) for i in range(n)]
    net_inc = [pbt[i]-tax_chg[i] for i in range(n)]

    # ── NWC & CF ──────────────────────────────────────────────────────────────
    ar   = [round(net_rev[i]*dso[i]/365) for i in range(n)]
    inv_ = [round(cogs[i]*dio[i]/365)    for i in range(n)]
    ap   = [round(cogs[i]*dpo[i]/365)    for i in range(n)]
    nwc  = [ar[i]+inv_[i]-ap[i]          for i in range(n)]
    d_nwc= [nwc[i]-nwc[i-1] if i>0 else 0 for i in range(n)]

    op_cf = [adj_ebitda[i]-tax_chg[i]-d_nwc[i] for i in range(n)]
    fcf   = [op_cf[i]-capex_t[i]               for i in range(n)]

    cash    = [max(0, fcf[i]*0.35)        for i in range(n)]
    net_dbt = [gross_debt[i]-cash[i]      for i in range(n)]

    # ── Company name ──────────────────────────────────────────────────────────
    company = "Business Plan"
    if "CONFIG" in sheets:
        for row in wb["CONFIG"].iter_rows(max_row=10, values_only=True):
            for v in row:
                s = str(v or "")
                if (len(s) > 4 and "BP GENERATOR" not in s
                        and "All model" not in s and "MASTER" not in s
                        and not s.startswith("1.")):
                    company = s
                    break
            if company != "Business Plan":
                break

    currency = "USD"
    for row in wb["ASSUMPTIONS"].iter_rows(max_row=4, values_only=True):
        for v in row:
            if v in ("USD","EUR","GBP","AED","BHD","SAR"):
                currency = str(v)

    # ── Row builder ───────────────────────────────────────────────────────────
    def row(label, vals, fmt="k0", indent=0, is_total=False, is_section=False):
        return {
            "label": label,
            "values": [round(v, 2) if isinstance(v, float) else v for v in vals],
            "fmt": fmt, "indent": indent,
            "is_total": is_total, "is_section": is_section,
        }

    def sec(label, cls=""):
        return row(label, [], fmt="section", is_section=True)

    # ── P&L section ───────────────────────────────────────────────────────────
    pl_rows = [
        sec("REVENUS"),
        row("Gross revenue",          gross_rev, indent=1),
        row("Net realisation",        net_rev,   is_total=True),
        sec("COÛTS VARIABLES"),
        row("Matières premières",     dm,   indent=1),
        row("Utilities",              util, indent=1),
        row("Packing",                pack, indent=1),
        row("Total COGS variable",    cogs, is_total=True),
        row("Marge brute",            gp,   is_total=True),
        row("  Marge brute %",
            [_div(gp[i],net_rev[i]) for i in range(n)], fmt="pct1", indent=1),
        sec("OPEX FIXES"),
        row("Personnel prod.",        sp,  indent=1),
        row("Personnel SG&A",         ss,  indent=1),
        row("Maintenance",            mnt, indent=1),
        row("Autres opex fixes",
            [ins[i]+rent[i]+it_s[i]+prof[i]+osga[i] for i in range(n)], indent=1),
        row("Total OpEx fixes",       fixed,     is_total=True),
        row("EBITDA (reported)",      ebitda,    is_total=True),
        row("EBITDA ajusté",          adj_ebitda,is_total=True),
        row("  Marge EBITDA adj. %",
            [_div(adj_ebitda[i],net_rev[i]) for i in range(n)], fmt="pct1", indent=1),
        sec("SOUS LA LIGNE"),
        row("D&A",                    dep,     indent=1),
        row("EBIT",                   ebit,    is_total=True),
        row("Intérêts (estimés)",     int_exp, indent=1),
        row("PBT",                    pbt,     is_total=True),
        row("Impôt",                  tax_chg, indent=1),
        row("RÉSULTAT NET",           net_inc, is_total=True),
        row("  Marge nette %",
            [_div(net_inc[i],net_rev[i]) for i in range(n)], fmt="pct1", indent=1),
    ]

    # ── KPIs section ──────────────────────────────────────────────────────────
    kpi_rows = [
        sec("VOLUME & PRIX"),
        row("Volume (MT)",     vol,       fmt="k0"),
        row("Prix moyen ($/MT)",price_mt, fmt="k1"),
        row("Capacité util. %",
            [_div(vol[i], capacity[i]) for i in range(n)], fmt="pct1"),
        sec("PROFITABILITÉ"),
        row("Marge brute %",
            [_div(gp[i],net_rev[i]) for i in range(n)], fmt="pct1"),
        row("EBITDA adj. ($k)", adj_ebitda),
        row("Marge EBITDA %",
            [_div(adj_ebitda[i],net_rev[i]) for i in range(n)], fmt="pct1"),
        row("EBITDA/MT ($/MT)",
            [round(adj_ebitda[i]*1000/vol[i]) if vol[i] else None for i in range(n)],
            fmt="k1"),
        row("Marge nette %",
            [_div(net_inc[i],net_rev[i]) for i in range(n)], fmt="pct1"),
        sec("LEVERAGE"),
        row("Dette brute ($k)",  gross_debt),
        row("Trésorerie ($k)",   cash),
        row("Dette nette ($k)",  net_dbt, is_total=True),
        row("Levier ND/EBITDA",
            [_div(net_dbt[i],adj_ebitda[i]) for i in range(n)], fmt="mult"),
        row("ICR (EBITDA/Int.)",
            [_div(adj_ebitda[i],int_exp[i]) if int_exp[i] else None
             for i in range(n)], fmt="mult"),
        sec("CASH FLOW"),
        row("Cash flow opérationnel", op_cf),
        row("Capex total",        capex_t,   indent=1),
        row("Free Cash Flow",     fcf,       is_total=True),
        row("FCF margin %",
            [_div(fcf[i],net_rev[i]) for i in range(n)], fmt="pct1"),
        sec("NWC"),
        row("DSO (jours)", dso, fmt="int0"),
        row("DPO (jours)", dpo, fmt="int0"),
        row("DIO (jours)", dio, fmt="int0"),
        row("NWC total",   nwc),
    ]

    # ── Debt section ──────────────────────────────────────────────────────────
    debt_section = []
    for t in debt_rows:
        debt_section.append(row(t["label"], t["balances"]))
    if gross_debt and any(v > 0 for v in gross_debt):
        debt_section.append(row("Total dette brute", gross_debt, is_total=True))
    if int_exp and any(v > 0 for v in int_exp):
        debt_section.append(row("Charge d'intérêts", int_exp))
    if repay and any(v > 0 for v in repay):
        debt_section.append(row("Remboursements", repay))

    return {
        "sheets":   sheets,
        "years":    years,
        "company":  company,
        "currency": currency,
        "sections": {
            "pl":   {"display_name": "Compte de résultat", "rows": pl_rows},
            "kpis": {"display_name": "KPIs & Ratios",      "rows": kpi_rows},
            "debt": {"display_name": "Structure de dette",  "rows": debt_section},
        },
    }
