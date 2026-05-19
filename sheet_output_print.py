from openpyxl.utils import get_column_letter
from styles import *

FIRST_DATA_COL = 5
LABEL_COL = 2
UNIT_COL  = 3
SRC_COL   = 4

def _col(idx):
    return get_column_letter(FIRST_DATA_COL + idx)

def _data_row(ws, row, label, formulas, font, num_fmt="k0",
              fill_=None, unit="$k", indent=0, row_ht=16, border=None):
    ws.row_dimensions[row].height = row_ht
    prefix = "    " * indent
    c = ws.cell(row=row, column=LABEL_COL, value=prefix + label)
    style_cell(c, font=font, align=AL["left"])
    c2 = ws.cell(row=row, column=UNIT_COL, value=unit)
    style_cell(c2, font=FONTS["italic"], align=AL["center"])
    for i, f in enumerate(formulas):
        c = ws.cell(row=row, column=FIRST_DATA_COL + i, value=f)
        kw = dict(font=font, align=AL["right"], num_fmt=NF[num_fmt])
        if fill_:  kw["fill_"] = fill_
        if border: kw["border"] = border
        style_cell(c, **kw)

def build_debt_schedule(wb, config, years, row_map):
    ws = wb.create_sheet("DEBT SCHEDULE")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    n = len(years)
    ccy = config["currency"]
    unit = f"{ccy}k"
    debt_cfg = config.get("debt", {})
    mechanics = config.get("debt_mechanics", {})

    # Column widths
    set_col_width(ws, "A", 2)
    set_col_width(ws, "B", 38)
    set_col_width(ws, "C", 8)
    set_col_width(ws, "D", 10)
    for i in range(n):
        set_col_width(ws, get_column_letter(FIRST_DATA_COL + i), 14)
    set_col_width(ws, get_column_letter(FIRST_DATA_COL + n), 28)

    row = 1

    # Title
    ws.row_dimensions[row].height = 34
    end = get_column_letter(FIRST_DATA_COL + n - 1)
    merge_and_style(ws, f"B{row}:{end}{row}",
                    f"{config['company_name']} — Debt Schedule & Interest",
                    FONTS["title"], fill(C["navy"]), AL["center"])
    row += 2

    # Period headers
    ws.row_dimensions[row].height = 18
    c = ws.cell(row=row, column=LABEL_COL, value="Tranche / Line item")
    style_cell(c, font=FONTS["header"], fill_=fill(C["debt_header"]), align=AL["left"])
    c2 = ws.cell(row=row, column=UNIT_COL, value="Units")
    style_cell(c2, font=FONTS["header"], fill_=fill(C["debt_header"]), align=AL["center"])
    c3 = ws.cell(row=row, column=SRC_COL, value="Rate")
    style_cell(c3, font=FONTS["header"], fill_=fill(C["debt_header"]), align=AL["center"])
    for i, yr in enumerate(years):
        c = ws.cell(row=row, column=FIRST_DATA_COL + i, value=yr["fy"])
        style_cell(c, font=FONTS["header"], fill_=fill(C["debt_header"]), align=AL["center"])
    row += 1

    ws.row_dimensions[row].height = 14
    for i, yr in enumerate(years):
        lbl = "Actuals" if yr["is_actual"] else "BP"
        c = ws.cell(row=row, column=FIRST_DATA_COL + i, value=lbl)
        clr = C["mid_grey"] if yr["is_actual"] else C["pale_blue"]
        style_cell(c, font=FONTS["note"], fill_=fill(clr), align=AL["center"])
    row += 1

    # ──────────────────────────────────────────────────────────────────────────
    # DEBT MAP — opening balances from CONFIG
    # ──────────────────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 20
    merge_and_style(ws, f"B{row}:{end}{row}", "A. DEBT MAP — OPENING BALANCES & TERMS",
                    FONTS["section"], fill(C["debt_header"]), AL["left"])
    row += 1

    # Parameters per active tranche
    active_tranches = []
    tranche_defs = [
        ("tla",  "Term Loan A — amortissable",    "Senior",  "quarterly"),
        ("tlb",  "Term Loan B — bullet",          "Senior",  "bullet"),
        ("ss",   "Super Senior (SSFA)",           "SS",      "quarterly"),
        ("rcf",  "Revolving Credit Facility",     "Senior",  "revolving"),
        ("cxf",  "Capex Facility",                "Senior",  "amortising"),
        ("mur",  "Islamic — Murabaha",            "Senior",  "quarterly"),
        ("ija",  "Islamic — Ijara / Wakala",      "Senior",  "quarterly"),
        ("mcf",  "Multi-currency Facility",       "Senior",  "amortising"),
        ("a1a2", "Tranche A1 / A2 split",         "Senior",  "quarterly"),
        ("mez",  "Mezzanine — cash pay",          "Mezz",    "bullet"),
        ("pik",  "PIK Toggle",                    "Mezz",    "bullet"),
        ("uni",  "Unitranche",                    "Senior",  "amortising"),
        ("sl",   "Second Lien",                   "Mezz",    "bullet"),
        ("shl",  "Shareholder Loan (SHL)",        "Sub",     "pik_only"),
        ("vnd",  "Vendor Loan / Seller Note",     "Sub",     "bullet"),
        ("cln",  "Convertible Loan Note",         "Sub",     "bullet"),
        ("pref", "Preferred Equity / Redeemable", "Sub",     "bullet"),
        ("ern",  "Earn-out",                      "Sub",     "contingent"),
        ("mud",  "Islamic Mudaraba / Musharaka",  "Sub",     "profit_share"),
        ("hyb",  "High Yield Bond",               "HYB",     "bullet"),
        ("ssn",  "Senior Secured Notes",          "Senior",  "bullet"),
        ("subn", "Subordinated Notes",            "Sub",     "bullet"),
        ("cvb",  "Convertible Bond",              "Sub",     "bullet"),
        ("epp",  "Euro PP / Schuldschein",        "Senior",  "amortising"),
        ("uspp", "USPP",                          "Senior",  "bullet"),
        ("suk",  "Sukuk",                         "Islamic", "bullet"),
        ("gsb",  "Green / Sustainability Bond",   "Senior",  "bullet"),
        ("pf",   "Project Finance — non-recourse","Senior",  "dscr"),
        ("bri",  "Bridge Loan",                   "Senior",  "bullet"),
        ("con",  "Construction Facility",         "Senior",  "amortising"),
        ("lea",  "Leasing Financier",             "Senior",  "amortising"),
        ("cbi",  "Crédit-bail Immobilier",        "Senior",  "amortising"),
        ("eca",  "ECA / State-backed",            "Senior",  "amortising"),
        ("ae",   "Amend & Extend",                "Senior",  "amortising"),
        ("dip",  "New Money (DIP-style)",         "SS",      "quarterly"),
        ("d2e",  "Debt-to-Equity Swap",           "Restr",   "n/a"),
        ("pikf", "PIK Forcé",                     "Restr",   "pik_only"),
        ("mor",  "Moratorium / Standstill",       "Restr",   "frozen"),
        ("cwv",  "Covenant Waiver / Reset",       "Restr",   "n/a"),
    ]

    for key, label, seniority, amort_type in tranche_defs:
        tranche_data = debt_cfg.get(key, {})
        if tranche_data.get("active", 0):
            active_tranches.append({
                "key": key, "label": label, "seniority": seniority,
                "amort_type": amort_type,
                "amount": tranche_data.get("amount", 0),
                "ccy":    tranche_data.get("currency", ccy),
                "tenor":  tranche_data.get("tenor_yrs", 5),
                "margin": tranche_data.get("margin_pct", 0.0325),
                "pik_margin": tranche_data.get("pik_margin_pct", 0.0),
            })

    if not active_tranches:
        # Fallback if no debt configured — show placeholder
        active_tranches = [{
            "key": "tla", "label": "Term Loan A", "seniority": "Senior",
            "amort_type": "quarterly", "amount": 150000, "ccy": ccy,
            "tenor": 7, "margin": 0.0325, "pik_margin": 0.0
        }]

    # Base rate assumption row in ASSUMPTIONS sheet
    base_rate_row = 50  # approximate row for base rate in ASSUMPTIONS
    base_rate_pct = mechanics.get("base_rate_pct", 0.045)
    upfront_fee   = mechanics.get("upfront_fee_pct", 0.015)
    cash_sweep    = mechanics.get("cash_sweep_pct", 0.75)
    sweep_active  = mechanics.get("cash_sweep_active", 1)
    lev_cov       = mechanics.get("lev_covenant", 5.0)
    icr_cov       = mechanics.get("icr_covenant", 2.0)

    # Debt map header row
    debt_map_start = row

    for t in active_tranches:
        ws.row_dimensions[row].height = 16
        # Opening balance row
        c = ws.cell(row=row, column=LABEL_COL, value=t["label"])
        style_cell(c, font=FONTS["label_bold"], align=AL["left"])

        c2 = ws.cell(row=row, column=UNIT_COL, value=t["ccy"]+"k")
        style_cell(c2, font=FONTS["italic"], align=AL["center"])

        rate_str = f"Base+{t['margin']*100:.2f}%"
        c3 = ws.cell(row=row, column=SRC_COL, value=rate_str)
        style_cell(c3, font=FONTS["note"], align=AL["center"])

        # Yr1 = opening balance (hardcoded input); subsequent = prior - repayment
        c_open = ws.cell(row=row, column=FIRST_DATA_COL, value=t["amount"])
        style_cell(c_open, font=FONTS["input"], fill_=fill(C["input_bg"]),
                   align=AL["right"], num_fmt=NF["k0"])

        # Amortisation schedule
        if t["amort_type"] in ("quarterly", "amortising"):
            amort_per_yr = t["amount"] / max(t["tenor"], 1)
            for i in range(1, n):
                prev_col = _col(i - 1)
                amort = amort_per_yr
                val = f"=MAX({prev_col}{row}-{amort:.0f},0)"
                c = ws.cell(row=row, column=FIRST_DATA_COL + i, value=val)
                style_cell(c, font=FONTS["formula"], align=AL["right"], num_fmt=NF["k0"])
        elif t["amort_type"] in ("bullet", "pik_only", "dscr", "frozen", "n/a",
                                 "revolving", "contingent", "profit_share"):
            for i in range(1, n):
                # Bullet: stays flat until maturity
                prev_col = _col(i - 1)
                c = ws.cell(row=row, column=FIRST_DATA_COL + i,
                            value=f"={prev_col}{row}")
                style_cell(c, font=FONTS["formula"], align=AL["right"], num_fmt=NF["k0"])

        t["balance_row"] = row
        row += 1

    # Total debt balance
    if active_tranches:
        bal_rows = [t["balance_row"] for t in active_tranches]
        for i in range(n):
            refs = "+".join([f"{_col(i)}{r}" for r in bal_rows])
            c = ws.cell(row=row, column=FIRST_DATA_COL + i, value=f"={refs}")
            style_cell(c, font=FONTS["total"], fill_=fill(C["total_fill"]),
                       align=AL["right"], num_fmt=NF["k0"],
                       border=BORDERS["thick"])
    c_lbl = ws.cell(row=row, column=LABEL_COL, value="Total gross debt")
    style_cell(c_lbl, font=FONTS["total"], fill_=fill(C["total_fill"]),
               align=AL["left"])
    row_map["ds_total_debt"] = row
    row += 2

    # ──────────────────────────────────────────────────────────────────────────
    # REPAYMENTS (annual)
    # ──────────────────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 20
    merge_and_style(ws, f"B{row}:{end}{row}", "B. ANNUAL DEBT REPAYMENTS",
                    FONTS["section"], fill(C["debt_header"]), AL["left"])
    row += 1

    for t in active_tranches:
        br = t["balance_row"]
        repay_fmls = []
        for i in range(n):
            if i == 0:
                repay_fmls.append(f"=0")
            else:
                repay_fmls.append(f"=MAX({_col(i-1)}{br}-{_col(i)}{br},0)")
        _data_row(ws, row, f"Repayment — {t['label']}", repay_fmls,
                  FONTS["label"], unit=t["ccy"]+"k", indent=1)
        t["repay_row"] = row
        row += 1

    if active_tranches:
        total_rep = [
            f"=SUM({_col(i)}{active_tranches[0]['repay_row']}:{_col(i)}{active_tranches[-1]['repay_row']})"
            for i in range(n)]
        _data_row(ws, row, "Total repayments", total_rep, FONTS["total"],
                  fill_=fill(C["total_fill"]), unit=unit, border=BORDERS["thick"])
        row_map["ds_total_repay"] = row
        row += 2

    # ──────────────────────────────────────────────────────────────────────────
    # INTEREST EXPENSE
    # ──────────────────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 20
    merge_and_style(ws, f"B{row}:{end}{row}", "C. INTEREST EXPENSE",
                    FONTS["section"], fill(C["debt_header"]), AL["left"])
    row += 1

    # Base rate
    _data_row(ws, row, "Base rate (SOFR / EURIBOR)", 
              [base_rate_pct] * n, FONTS["input"],
              fill_=fill(C["input_bg"]), unit="%", num_fmt="pct2", indent=1)
    base_rate_data_row = row; row += 1

    _data_row(ws, row, "Base rate floor",
              [mechanics.get("base_rate_floor", 0.0)] * n, FONTS["input"],
              fill_=fill(C["input_bg"]), unit="%", num_fmt="pct2", indent=1)
    floor_row = row; row += 1

    for t in active_tranches:
        br = t["balance_row"]
        is_pik = t["amort_type"] in ("pik_only",)
        margin = t["margin"]
        pik_m  = t["pik_margin"]

        # Cash interest = avg balance × MAX(base+margin, floor)
        if not is_pik:
            int_fmls = [
                f"=({_col(i)}{br}+(IF({i}>0,{_col(i-1)}{br},{_col(i)}{br})))/2"
                f"*MAX({_col(i)}{base_rate_data_row}+{margin},{_col(i)}{floor_row})"
                if i > 0 else
                f"={_col(i)}{br}*MAX({_col(i)}{base_rate_data_row}+{margin},{_col(i)}{floor_row})"
                for i in range(n)]
            _data_row(ws, row, f"Interest — {t['label']} (cash)",
                      int_fmls, FONTS["formula"], unit=unit, indent=1)
            t["int_cash_row"] = row; row += 1

        if pik_m > 0 or is_pik:
            pik_fmls = [
                f"={_col(i)}{br}*{pik_m}" for i in range(n)]
            _data_row(ws, row, f"Interest — {t['label']} (PIK)",
                      pik_fmls, FONTS["formula"], unit=unit, indent=1)
            t["int_pik_row"] = row; row += 1

    # Total cash interest
    cash_int_rows = [t.get("int_cash_row") for t in active_tranches if t.get("int_cash_row")]
    if cash_int_rows:
        tot_cash = [f"=SUM({','.join([f'{_col(i)}{r}' for r in cash_int_rows])})"
                    for i in range(n)]
        _data_row(ws, row, "Total cash interest", tot_cash, FONTS["total"],
                  fill_=fill(C["total_fill"]), unit=unit, border=BORDERS["thick"])
        row_map["ds_cash_interest"] = row; row += 1
    else:
        row_map["ds_cash_interest"] = row

    # Total PIK interest
    pik_int_rows = [t.get("int_pik_row") for t in active_tranches if t.get("int_pik_row")]
    if pik_int_rows:
        tot_pik = [f"=SUM({','.join([f'{_col(i)}{r}' for r in pik_int_rows])})"
                   for i in range(n)]
        _data_row(ws, row, "Total PIK interest", tot_pik, FONTS["subtotal"],
                  fill_=fill(C["subtot_fill"]), unit=unit)
        row_map["ds_pik_interest"] = row; row += 1
    else:
        row_map["ds_pik_interest"] = row

    # Upfront fees (amortised)
    total_debt_amt = sum(t["amount"] for t in active_tranches)
    avg_tenor = sum(t["tenor"] * t["amount"] for t in active_tranches) / max(total_debt_amt, 1)
    fee_annual = (total_debt_amt * upfront_fee) / max(avg_tenor, 1)
    fee_fmls = [str(round(fee_annual)) for _ in range(n)]
    _data_row(ws, row, "Upfront fees — amortised", fee_fmls, FONTS["formula"],
              unit=unit, indent=1)
    row_map["ds_fees_amort"] = row; row += 1

    # Rows 50/51/52 — referenced from P&L sheet
    # We ensure these rows are exactly where the P&L expects them
    row_map["ds_int_cash_pl_ref"]  = row_map.get("ds_cash_interest", row)
    row_map["ds_int_pik_pl_ref"]   = row_map.get("ds_pik_interest", row)
    row_map["ds_fees_pl_ref"]      = row_map.get("ds_fees_amort", row)

    row += 2

    # ──────────────────────────────────────────────────────────────────────────
    # CASH SWEEP
    # ──────────────────────────────────────────────────────────────────────────
    if sweep_active:
        ws.row_dimensions[row].height = 20
        merge_and_style(ws, f"B{row}:{end}{row}", "D. CASH SWEEP",
                        FONTS["section"], fill(C["debt_header"]), AL["left"])
        row += 1

        sweep_pct_fmls = [str(cash_sweep)] * n
        _data_row(ws, row, "Cash sweep % (excess CF)", sweep_pct_fmls,
                  FONTS["input"], fill_=fill(C["input_bg"]),
                  unit="%", num_fmt="pct1", indent=1)
        sweep_pct_row = row; row += 1

        # Excess cash = FCF (from CF sheet row ~45) × sweep %
        sweep_fmls = [
            f"=IFERROR(MAX('CASH FLOW'!{_col(i)}45,0)*{_col(i)}{sweep_pct_row},0)"
            for i in range(n)]
        _data_row(ws, row, "Mandatory cash sweep", sweep_fmls, FONTS["formula"],
                  unit=unit, indent=1)
        row_map["ds_cash_sweep"] = row; row += 2

    # ──────────────────────────────────────────────────────────────────────────
    # COVENANT TRACKING
    # ──────────────────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 20
    merge_and_style(ws, f"B{row}:{end}{row}", "E. COVENANT TRACKING",
                    FONTS["section"], fill(C["debt_header"]), AL["left"])
    row += 1

    # Net leverage = Net debt / Adj EBITDA
    net_lev_fmls = [
        f"=IFERROR({_col(i)}{row_map.get('ds_total_debt',5)}/"
        f"'P&L'!{_col(i)}{row_map.get('pl_ebitda_adj',50)},\"-\")"
        for i in range(n)]
    _data_row(ws, row, "Net leverage (Net debt / Adj EBITDA)", net_lev_fmls,
              FONTS["formula"], unit="x", num_fmt="mult")
    cov_lev_row = row; row += 1

    cov_lev_limit = [str(lev_cov)] * n
    _data_row(ws, row, "Covenant — max leverage", cov_lev_limit,
              FONTS["input"], fill_=fill(C["input_bg"]), unit="x", num_fmt="mult", indent=1)
    row_map["ds_lev_limit"] = row; row += 1

    # Headroom
    hdroom_lev = [
        f"=IFERROR({_col(i)}{row_map['ds_lev_limit']}-{_col(i)}{cov_lev_row},\"-\")"
        for i in range(n)]
    lev_colors = []
    _data_row(ws, row, "Headroom — leverage", hdroom_lev, FONTS["formula"],
              unit="x", num_fmt="mult", indent=1)
    row += 1

    # ICR = Adj EBITDA / Net interest
    icr_fmls = [
        f"=IFERROR('P&L'!{_col(i)}{row_map.get('pl_ebitda_adj',50)}/"
        f"'P&L'!{_col(i)}{row_map.get('pl_net_interest',60)},\"-\")"
        for i in range(n)]
    _data_row(ws, row, "Interest cover (Adj EBITDA / Net interest)", icr_fmls,
              FONTS["formula"], unit="x", num_fmt="mult")
    cov_icr_row = row; row += 1

    icr_min = [str(icr_cov)] * n
    _data_row(ws, row, "Covenant — min ICR", icr_min,
              FONTS["input"], fill_=fill(C["input_bg"]), unit="x", num_fmt="mult", indent=1)
    row_map["ds_icr_limit"] = row; row += 1

    hdroom_icr = [
        f"=IFERROR({_col(i)}{cov_icr_row}-{_col(i)}{row_map['ds_icr_limit']},\"-\")"
        for i in range(n)]
    _data_row(ws, row, "Headroom — ICR", hdroom_icr, FONTS["formula"],
              unit="x", num_fmt="mult", indent=1)
    row += 2

    # Covenant breach flags
    ws.row_dimensions[row].height = 18
    merge_and_style(ws, f"B{row}:{end}{row}", "Covenant breach flags (red = breach)",
                    FONTS["subheader"], fill(C["debt_sub"]), AL["left"])
    row += 1

    for i in range(n):
        c = ws.cell(row=row, column=FIRST_DATA_COL + i,
                    value=f"=IF({_col(i)}{cov_lev_row}<={_col(i)}{row_map['ds_lev_limit']},\"OK\",\"BREACH\")")
        ok_fill   = fill(C["check_ok"])
        err_fill  = fill(C["check_err"])
        style_cell(c, font=FONTS["formula"], align=AL["center"])
    c_lbl = ws.cell(row=row, column=LABEL_COL, value="Leverage covenant")
    style_cell(c_lbl, font=FONTS["label"], align=AL["left"])
    row += 1

    for i in range(n):
        c = ws.cell(row=row, column=FIRST_DATA_COL + i,
                    value=f"=IF({_col(i)}{cov_icr_row}>={_col(i)}{row_map['ds_icr_limit']},\"OK\",\"BREACH\")")
        style_cell(c, font=FONTS["formula"], align=AL["center"])
    c_lbl = ws.cell(row=row, column=LABEL_COL, value="ICR covenant")
    style_cell(c_lbl, font=FONTS["label"], align=AL["left"])
    row += 2

    # ──────────────────────────────────────────────────────────────────────────
    # WATERFALL (simplified annual)
    # ──────────────────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 20
    merge_and_style(ws, f"B{row}:{end}{row}",
                    "F. CASH WATERFALL — ORDER OF PRIORITY",
                    FONTS["section"], fill(C["debt_header"]), AL["left"])
    row += 1

    wf_items = [
        ("Operating cash flow (pre-debt service)",  f"'CASH FLOW'!{'{col}'}{45}"),
        ("Less: mandatory debt repayment",          f"'DEBT SCHEDULE'!{'{col}'}{row_map.get('ds_total_repay',10)}"),
        ("Less: cash interest (total)",             f"'DEBT SCHEDULE'!{'{col}'}{row_map.get('ds_cash_interest',10)}"),
        ("Less: mandatory cash sweep",              f"'DEBT SCHEDULE'!{'{col}'}{row_map.get('ds_cash_sweep',10)}" if sweep_active else "0"),
        ("= Free cash flow to equity",              None),
    ]

    cumulative_rows = []
    for wf_label, wf_ref in wf_items:
        ws.row_dimensions[row].height = 17
        if wf_ref is None:
            cum_fmls = []
            for i in range(n):
                refs = "-".join([f"{_col(i)}{r}" for r in cumulative_rows[1:]])
                cum_fmls.append(f"={_col(i)}{cumulative_rows[0]}-{refs}" if refs else f"={_col(i)}{cumulative_rows[0]}")
            _data_row(ws, row, wf_label, cum_fmls, FONTS["total"],
                      fill_=fill(C["total_fill"]), unit=unit, border=BORDERS["thick"])
        else:
            fmls = [wf_ref.replace("{col}", _col(i)) for i in range(n)]
            _data_row(ws, row, wf_label, fmls, FONTS["formula"], unit=unit, indent=1)
        cumulative_rows.append(row)
        row += 1

    ws.sheet_properties.tabColor = "7B0000"
    return ws
