"""
Financial Data Extractor
Reads Excel / CSV / PDF and maps to BP Generator assumptions format.
Returns a structured dict compatible with the BP build engine.
"""
import re
import json
import logging
from pathlib import Path
from typing import Any

import pandas as pd
import pdfplumber
import openpyxl

logger = logging.getLogger(__name__)

# ── Keyword maps for line-item recognition ───────────────────────────────────
# Each entry: (canonical_key, [list of keywords that identify it])
LINE_MAPS = {
    # P&L
    "revenue":          ["revenue","turnover","chiffre d'affaires","ca","net sales","sales"],
    "gross_profit":     ["gross profit","marge brute","résultat brut"],
    "ebitda":           ["ebitda","excédent brut","ebe"],
    "ebit":             ["ebit","operating profit","résultat opérationnel","résultat d'exploitation"],
    "net_income":       ["net income","net profit","résultat net","profit net","bénéfice net"],
    "depreciation":     ["depreciation","amortisation","amortissement","d&a","d & a"],
    "interest_expense": ["interest expense","charge financière","frais financiers","intérêts"],
    "tax":              ["income tax","impôt","tax charge","corporation tax"],
    # Balance Sheet
    "total_assets":     ["total assets","total actif","actif total"],
    "total_equity":     ["total equity","capitaux propres","equity"],
    "gross_debt":       ["gross debt","financial debt","dette financière","borrowings","loans"],
    "cash":             ["cash","trésorerie","liquidités","cash and equivalents"],
    "net_debt":         ["net debt","dette nette"],
    "trade_receivables":["trade receivables","créances clients","accounts receivable","debtors"],
    "inventories":      ["inventories","stocks","inventory"],
    "trade_payables":   ["trade payables","dettes fournisseurs","accounts payable","creditors"],
    # CF
    "operating_cf":     ["operating cash","cash from operations","flux opérationnel","cfo"],
    "capex":            ["capex","capital expenditure","investissements","acquisition of ppe"],
    "fcf":              ["free cash flow","flux de trésorerie libre","fcf"],
    # Industrial KPIs
    "volume_mt":        ["volume","volume mt","tonnage","quantity sold"],
    "price_per_mt":     ["price per mt","prix unitaire","average price","asp"],
}

def _normalize(text: str) -> str:
    """Lowercase, strip accents-lite, remove special chars for matching."""
    if not isinstance(text, str):
        return ""
    t = text.lower().strip()
    t = t.replace("'", "'").replace("–", "-").replace("—", "-")
    return re.sub(r"[^\w\s&/'-]", " ", t)

def _match_line(label: str) -> str | None:
    """Return canonical key if label matches any keyword."""
    norm = _normalize(label)
    for key, keywords in LINE_MAPS.items():
        for kw in keywords:
            if kw in norm:
                return key
    return None

def _to_float(val: Any) -> float | None:
    """Convert various cell values to float."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if not pd.isna(val) else None
    s = str(val).replace(",", "").replace(" ", "").replace("(", "-").replace(")", "")
    s = re.sub(r"[^\d.\-]", "", s)
    try:
        return float(s) if s and s != "-" else None
    except ValueError:
        return None

def _detect_years(columns) -> list[int]:
    """Extract year integers from column headers."""
    years = []
    for col in columns:
        m = re.search(r"(20\d{2}|19\d{2})", str(col))
        if m:
            years.append(int(m.group(1)))
    return sorted(set(years))

# ════════════════════════════════════════════════════════════════════════════
# EXCEL EXTRACTOR
# ════════════════════════════════════════════════════════════════════════════
def extract_excel(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    results = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find year headers
        year_cols = {}  # year -> col_index
        for row in rows[:10]:
            for ci, cell in enumerate(row):
                m = re.search(r"(20\d{2})", str(cell or ""))
                if m:
                    yr = int(m.group(1))
                    if yr not in year_cols:
                        year_cols[yr] = ci

        if not year_cols:
            continue

        # Scan rows for financial line items
        for row in rows:
            label = str(row[0] or "")
            key = _match_line(label)
            if not key:
                continue
            if key not in results:
                results[key] = {}
            for yr, ci in year_cols.items():
                if ci < len(row):
                    val = _to_float(row[ci])
                    if val is not None and yr not in results[key]:
                        results[key][yr] = val

    return results

# ════════════════════════════════════════════════════════════════════════════
# CSV EXTRACTOR
# ════════════════════════════════════════════════════════════════════════════
def extract_csv(path: str) -> dict:
    results = {}

    # Try common separators
    for sep in [",", ";", "\t", "|"]:
        try:
            df = pd.read_csv(path, sep=sep, encoding="utf-8", on_bad_lines="skip")
            if df.shape[1] < 2:
                continue

            years = _detect_years(df.columns[1:])
            if not years:
                # Try first row as header
                df = pd.read_csv(path, sep=sep, header=1, encoding="utf-8", on_bad_lines="skip")
                years = _detect_years(df.columns[1:])

            if not years:
                continue

            for _, row in df.iterrows():
                label = str(row.iloc[0])
                key = _match_line(label)
                if not key:
                    continue
                if key not in results:
                    results[key] = {}
                for ci, col in enumerate(df.columns[1:]):
                    m = re.search(r"(20\d{2})", str(col))
                    if m:
                        yr = int(m.group(1))
                        val = _to_float(row.iloc[ci + 1])
                        if val is not None and yr not in results[key]:
                            results[key][yr] = val

            if results:
                break
        except Exception:
            continue

    return results

# ════════════════════════════════════════════════════════════════════════════
# PDF EXTRACTOR
# ════════════════════════════════════════════════════════════════════════════
def extract_pdf(path: str) -> dict:
    results = {}

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            # Try structured tables first
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 2:
                    continue

                # Find year columns in header row
                header = table[0] or []
                year_cols = {}
                for ci, cell in enumerate(header):
                    m = re.search(r"(20\d{2})", str(cell or ""))
                    if m:
                        year_cols[int(m.group(1))] = ci

                if not year_cols:
                    continue

                for row in table[1:]:
                    if not row:
                        continue
                    label = str(row[0] or "")
                    key = _match_line(label)
                    if not key:
                        continue
                    if key not in results:
                        results[key] = {}
                    for yr, ci in year_cols.items():
                        if ci < len(row):
                            val = _to_float(row[ci])
                            if val is not None and yr not in results[key]:
                                results[key][yr] = val

            # Fallback: raw text parsing
            if not results:
                text = page.extract_text() or ""
                lines = text.split("\n")
                years_found = []
                for line in lines:
                    yrs = re.findall(r"20\d{2}", line)
                    if yrs:
                        years_found = [int(y) for y in yrs]
                        break

                for line in lines:
                    key = _match_line(line)
                    if not key:
                        continue
                    nums = re.findall(r"-?\(?\d[\d,. ]+\)?", line)
                    vals = [_to_float(n) for n in nums if _to_float(n) is not None]
                    if vals and years_found:
                        if key not in results:
                            results[key] = {}
                        for i, yr in enumerate(years_found[:len(vals)]):
                            if yr not in results[key]:
                                results[key][yr] = vals[i]

    return results

# ════════════════════════════════════════════════════════════════════════════
# MAIN EXTRACTOR — auto-detects format
# ════════════════════════════════════════════════════════════════════════════
def extract_financials(path: str) -> dict:
    """
    Auto-detect file format and extract financial data.
    Returns raw dict: {canonical_key: {year: value}}
    """
    p = Path(path)
    ext = p.suffix.lower()

    if ext in [".xlsx", ".xls", ".xlsm"]:
        return extract_excel(path)
    elif ext in [".csv", ".txt", ".tsv"]:
        return extract_csv(path)
    elif ext == ".pdf":
        return extract_pdf(path)
    else:
        raise ValueError(f"Format non supporté : {ext}")

# ════════════════════════════════════════════════════════════════════════════
# MAPPER — converts raw extracted data to BP config actuals format
# ════════════════════════════════════════════════════════════════════════════
def map_to_bp_actuals(raw: dict, n_proj_years: int = 7) -> dict:
    """
    Maps extracted financials to BP actuals overlay.
    Returns dict with years as keys and financial line items as values.
    """
    # Gather all years found
    all_years = set()
    for v in raw.values():
        all_years.update(v.keys())
    hist_years = sorted(all_years)

    actuals = {}
    for yr in hist_years:
        actuals[yr] = {}
        for key in LINE_MAPS.keys():
            if key in raw and yr in raw[key]:
                actuals[yr][key] = raw[key][yr]

    # Derive missing items where possible
    for yr in hist_years:
        d = actuals[yr]
        # Net debt = gross debt - cash
        if "net_debt" not in d and "gross_debt" in d and "cash" in d:
            d["net_debt"] = d["gross_debt"] - d["cash"]
        # EBITDA = EBIT + D&A
        if "ebitda" not in d and "ebit" in d and "depreciation" in d:
            d["ebitda"] = d["ebit"] + d["depreciation"]
        # EBIT = EBITDA - D&A
        if "ebit" not in d and "ebitda" in d and "depreciation" in d:
            d["ebit"] = d["ebitda"] - d["depreciation"]
        # Revenue growth rate
        if yr > hist_years[0] and "revenue" in d:
            prev = actuals.get(yr-1, {}).get("revenue")
            if prev and prev != 0:
                d["revenue_growth"] = (d["revenue"] - prev) / abs(prev)
        # EBITDA margin
        if "ebitda" in d and "revenue" in d and d["revenue"] != 0:
            d["ebitda_margin"] = d["ebitda"] / d["revenue"]
        # Net margin
        if "net_income" in d and "revenue" in d and d["revenue"] != 0:
            d["net_margin"] = d["net_income"] / d["revenue"]
        # Leverage
        if "net_debt" in d and "ebitda" in d and d["ebitda"] != 0:
            d["net_leverage"] = d["net_debt"] / d["ebitda"]

    return {
        "hist_years":   hist_years,
        "actuals":      actuals,
        "raw":          raw,
        "data_quality": _assess_quality(raw, hist_years),
    }

def _assess_quality(raw: dict, years: list) -> dict:
    """Score data quality — which key line items are present."""
    core    = ["revenue", "ebitda", "net_income"]
    extra   = ["gross_profit", "ebit", "depreciation", "interest_expense",
                "total_assets", "gross_debt", "cash", "capex"]
    kpi     = ["volume_mt", "price_per_mt", "operating_cf"]

    found   = set(raw.keys())
    score   = 0
    issues  = []
    present = {}

    for k in core:
        if k in found:
            score += 30; present[k] = True
        else:
            issues.append(f"Manquant (critique) : {k}")
            present[k] = False

    for k in extra:
        if k in found:
            score += 5; present[k] = True
        else:
            present[k] = False

    for k in kpi:
        if k in found:
            score += 3; present[k] = True
        else:
            present[k] = False

    score = min(score, 100)

    return {
        "score":   score,
        "years":   years,
        "present": present,
        "issues":  issues,
        "label":   "Excellent" if score>=85 else "Bon" if score>=60 else "Partiel" if score>=30 else "Insuffisant",
    }

# ════════════════════════════════════════════════════════════════════════════
# BUILD PROJECTION ASSUMPTIONS from actuals
# ════════════════════════════════════════════════════════════════════════════
def build_projections_from_actuals(mapped: dict, n_years: int = 7) -> dict:
    """
    Uses last 2-3 years of actuals to derive projection assumptions.
    Returns a partial BP config dict with revenue/cost/nwc drivers.
    """
    actuals = mapped["actuals"]
    years   = mapped["hist_years"]
    if not years:
        return {}

    def avg(key, last_n=3):
        vals = [actuals[y].get(key) for y in years[-last_n:] if actuals[y].get(key) is not None]
        return sum(vals)/len(vals) if vals else None

    def last(key):
        for y in reversed(years):
            v = actuals[y].get(key)
            if v is not None:
                return v
        return None

    def growth_rate(key, last_n=3):
        vals = [(y, actuals[y][key]) for y in years if actuals[y].get(key) is not None]
        if len(vals) < 2:
            return 0.03
        rates = []
        for i in range(1, min(last_n, len(vals))):
            prev = vals[-i-1][1]; curr = vals[-i][1]
            if prev and prev != 0:
                rates.append((curr - prev) / abs(prev))
        return sum(rates)/len(rates) if rates else 0.03

    last_rev   = last("revenue")     or 0
    rev_growth = growth_rate("revenue")
    ebitda_mgn = avg("ebitda_margin") or 0.15
    net_mgn    = avg("net_margin")    or 0.05
    last_capex = last("capex")        or 0

    # Project revenue
    proj_rev = []
    r = last_rev
    for _ in range(n_years):
        r = r * (1 + rev_growth)
        proj_rev.append(round(r))

    # Project EBITDA (margin-based)
    proj_ebitda = [round(r * ebitda_mgn) for r in proj_rev]

    # Derive fixed opex from EBITDA and gross profit if available
    gp_rev_ratio = avg("gross_profit") / last_rev if last_rev and avg("gross_profit") else 0.35
    proj_gp = [round(r * gp_rev_ratio) for r in proj_rev]

    # Fixed opex = GP - EBITDA
    proj_fixed_opex = [max(0, gp - eb) for gp, eb in zip(proj_gp, proj_ebitda)]

    # NWC days — derive from last year if AR/AP/Inv available
    last_ar  = last("trade_receivables")
    last_ap  = last("trade_payables")
    last_inv = last("inventories")
    dso = round(last_ar  / last_rev  * 365) if last_ar  and last_rev  else 45
    dpo = round(last_ap  / (last_rev * (1-gp_rev_ratio)) * 365) if last_ap and last_rev else 55
    dio = round(last_inv / (last_rev * (1-gp_rev_ratio)) * 365) if last_inv and last_rev else 35
    dso = max(10, min(dso, 180))
    dpo = max(10, min(dpo, 180))
    dio = max(5,  min(dio, 120))

    return {
        "actuals_overlay": actuals,
        "hist_years":      years,
        "proj_assumptions": {
            "revenue":        proj_rev,
            "ebitda":         proj_ebitda,
            "revenue_growth": [round(rev_growth, 4)] * n_years,
            "ebitda_margin":  [round(ebitda_mgn, 4)] * n_years,
            "net_margin":     [round(net_mgn, 4)]    * n_years,
            "fixed_opex":     proj_fixed_opex,
            "capex_maint":    [round(last_capex * 0.6)] * n_years,
            "capex_expan":    [round(last_capex * 0.4)] * n_years,
            "dso":            [dso] * n_years,
            "dpo":            [dpo] * n_years,
            "dio":            [dio] * n_years,
        },
        "last_actuals": {
            "revenue":    last_rev,
            "ebitda":     last("ebitda"),
            "net_income": last("net_income"),
            "gross_debt": last("gross_debt"),
            "cash":       last("cash"),
            "net_debt":   last("net_debt"),
            "capex":      last_capex,
        },
        "data_quality": mapped["data_quality"],
    }
