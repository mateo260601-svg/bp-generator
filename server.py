from openpyxl.utils import get_column_letter
from styles import *
from periods import build_periods, col_for_period

FIRST_DATA_COL = 5   # Column E onwards = data
LABEL_COL      = 2   # B
UNIT_COL       = 3   # C
SOURCE_COL     = 4   # D

def _write_period_headers(ws, row, years, monthly_periods, annual_only=False):
    """Write year/month header rows. Returns next row."""
    # Annual headers
    ws.row_dimensions[row].height = 18
    c = ws.cell(row=row, column=LABEL_COL, value="Fiscal Year")
    style_cell(c, font=FONTS["header"], fill_=fill(C["assump_hdr"]), align=AL["center"])
    for i, yr in enumerate(years):
        col = get_column_letter(FIRST_DATA_COL + i)
        c = ws.cell(row=row, column=FIRST_DATA_COL + i, value=yr["fy"])
        style_cell(c, font=FONTS["header"], fill_=fill(C["assump_hdr"]), align=AL["center"])
    row += 1

    if not annual_only:
        # Actuals / BP flag row
        ws.row_dimensions[row].height = 14
        for i, yr in enumerate(years):
            col_idx = FIRST_DATA_COL + i
            lbl = "Actuals" if yr["is_actual"] else "BP"
            c = ws.cell(row=row, column=col_idx, value=lbl)
            clr = C["mid_grey"] if yr["is_actual"] else C["pale_blue"]
            style_cell(c, font=FONTS["note"], fill_=fill(clr), align=AL["center"])
        row += 1

    return row

def _section_header(ws, row, title, color_key, n_cols):
    ws.row_dimensions[row].height = 20
    end_col = get_column_letter(FIRST_DATA_COL + n_cols - 1)
    merge_and_style(ws, f"B{row}:{end_col}{row}", title,
                    FONTS["section"], fill(C[color_key]), AL["left"])
    row += 1
    return row

def _sub_header(ws, row, title, color_key, n_cols):
    ws.row_dimensions[row].height = 16
    end_col = get_column_letter(FIRST_DATA_COL + n_cols - 1)
    merge_and_style(ws, f"B{row}:{end_col}{row}", title,
                    FONTS["subheader"], fill(C[color_key]), AL["left"])
    row += 1
    return row

def _assumption_row(ws, row, label, unit, source, values, fmt="k0",
                    is_input=True, note=None):
    """Write one assumption row with annual values."""
    ws.row_dimensions[row].height = 16

    c_lbl = ws.cell(row=row, column=LABEL_COL, value=label)
    style_cell(c_lbl, font=FONTS["label"], align=AL["left"])

    c_unit = ws.cell(row=row, column=UNIT_COL, value=unit)
    style_cell(c_unit, font=FONTS["italic"], align=AL["center"])

    c_src = ws.cell(row=row, column=SOURCE_COL, value=source)
    style_cell(c_src, font=FONTS["italic"], align=AL["center"])

    for i, val in enumerate(values):
        c = ws.cell(row=row, column=FIRST_DATA_COL + i, value=val)
        if is_input:
            style_cell(c, font=FONTS["input"], fill_=fill(C["input_bg"]),
                       align=AL["right"], num_fmt=NF[fmt])
        else:
            style_cell(c, font=FONTS["formula"], align=AL["right"], num_fmt=NF[fmt])

    if note:
        last_col = FIRST_DATA_COL + len(values)
        c_note = ws.cell(row=row, column=last_col, value=note)
        style_cell(c_note, font=FONTS["note"], align=AL["left"])

    return row + 1

def build_assumptions(wb, config, years, monthly_periods):
    ws = wb.create_sheet("ASSUMPTIONS")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    n = len(years)

    # Column widths
    set_col_width(ws, "A", 2)
    set_col_width(ws, "B", 36)
    set_col_width(ws, "C", 10)
    set_col_width(ws, "D", 12)
    for i in range(n):
        set_col_width(ws, get_column_letter(FIRST_DATA_COL + i), 14)
    set_col_width(ws, get_column_letter(FIRST_DATA_COL + n), 30)

    row = 1

    # Title
    ws.row_dimensions[row].height = 34
    end = get_column_letter(FIRST_DATA_COL + n - 1)
    merge_and_style(ws, f"B{row}:{end}{row}",
                    f"{config['company_name']} — Assumptions & Drivers",
                    FONTS["title"], fill(C["navy"]), AL["center"])
    row += 2

    # Period headers
    row = _write_period_headers(ws, row, years, monthly_periods, annual_only=True)
    header_row = row - 2  # save for later reference

    # ────────────────────────────────────────────────────────────────────────
    # REVENUE DRIVERS
    # ────────────────────────────────────────────────────────────────────────
    row = _section_header(ws, row, "1. REVENUE DRIVERS", "pl_header", n + 3)
    row = _sub_header(ws, row, "1.1  Volume & pricing (industrial)", "pl_sub", n + 3)

    rev = config.get("revenue", {})
    base_vol  = rev.get("base_volume",   [180000]*n)
    vol_gr    = rev.get("volume_growth", [0.03]*n)
    price_mt  = rev.get("price_per_mt",  [1050]*n)
    freight   = rev.get("freight_mt",    [35]*n)
    commission= rev.get("commission_pct",[0.01]*n)
    capacity  = rev.get("capacity_mt",   [220000]*n)
    util_pct  = [v/c if c else 0 for v,c in zip(base_vol, capacity)]

    row = _assumption_row(ws, row, "Installed capacity",    "MT",  "Config",  capacity,   "k0")
    row = _assumption_row(ws, row, "Sales volume",          "MT",  "Input",   base_vol,   "k0")
    row = _assumption_row(ws, row, "Capacity utilisation",  "%",   "Calc",    util_pct,   "pct1", is_input=False)
    row = _assumption_row(ws, row, "Volume growth YoY",     "%",   "Input",   vol_gr,     "pct1")
    row = _assumption_row(ws, row, "Average sales price",   "$/MT","Input",   price_mt,   "k1")
    row = _assumption_row(ws, row, "Freight & forwarding",  "$/MT","Input",   freight,    "k1")
    row = _assumption_row(ws, row, "Sales commission",      "%",   "Input",   commission, "pct2",
                          note="% of gross revenue")
    row += 1

    row = _sub_header(ws, row, "1.2  Net realisation build", "pl_sub", n + 3)
    # Calculated: Gross revenue, net revenue
    gross_rev = [v * p / 1000 for v, p in zip(base_vol, price_mt)]  # $k
    net_rev   = [g - v * f / 1000 - g * cm
                 for g, v, f, cm in zip(gross_rev, base_vol, freight, commission)]
    row = _assumption_row(ws, row, "Gross revenue",          f"{config['currency']}k", "Calc",
                          gross_rev, "k0", is_input=False)
    row = _assumption_row(ws, row, "Less: freight & forwarding", f"{config['currency']}k", "Calc",
                          [v * f / 1000 for v, f in zip(base_vol, freight)], "k0", is_input=False)
    row = _assumption_row(ws, row, "Less: sales commission", f"{config['currency']}k", "Calc",
                          [g * cm for g, cm in zip(gross_rev, commission)], "k0", is_input=False)
    row = _assumption_row(ws, row, "Net realisation",        f"{config['currency']}k", "Calc",
                          net_rev, "k0", is_input=False)
    row += 1

    # ────────────────────────────────────────────────────────────────────────
    # COST DRIVERS
    # ────────────────────────────────────────────────────────────────────────
    row = _section_header(ws, row, "2. COST DRIVERS", "pl_header", n + 3)
    row = _sub_header(ws, row, "2.1  Direct / variable costs", "pl_sub", n + 3)

    cst = config.get("costs", {})
    row = _assumption_row(ws, row, "Direct materials",       "$/MT","Input",
                          cst.get("direct_mat_mt",   [520]*n), "k1")
    row = _assumption_row(ws, row, "Direct materials growth","%",  "Input",
                          cst.get("direct_mat_gr",   [0.02]*n), "pct1")
    row = _assumption_row(ws, row, "Utilities",              "$/MT","Input",
                          cst.get("utilities_mt",    [45]*n),  "k1")
    row = _assumption_row(ws, row, "Packing costs",          "$/MT","Input",
                          cst.get("packing_mt",      [18]*n),  "k1")
    row = _assumption_row(ws, row, "Variable OpEx — other",  "$/MT","Input",
                          cst.get("var_opex_other_mt",[12]*n), "k1")
    row += 1

    row = _sub_header(ws, row, "2.2  Fixed costs (annual)", "pl_sub", n + 3)
    row = _assumption_row(ws, row, "Personnel — production", f"{config['currency']}k","Input",
                          cst.get("staff_prod",   [8500]*n),  "k0")
    row = _assumption_row(ws, row, "Personnel — SG&A",       f"{config['currency']}k","Input",
                          cst.get("staff_sga",    [3200]*n),  "k0")
    row = _assumption_row(ws, row, "Headcount growth",       "%",   "Input",
                          cst.get("headcount_gr", [0.02]*n),  "pct1")
    row = _assumption_row(ws, row, "Maintenance & repairs",  f"{config['currency']}k","Input",
                          cst.get("maintenance",  [2400]*n),  "k0")
    row = _assumption_row(ws, row, "Insurance",              f"{config['currency']}k","Input",
                          cst.get("insurance",    [800]*n),   "k0")
    row = _assumption_row(ws, row, "Rent / site costs",      f"{config['currency']}k","Input",
                          cst.get("rent",         [600]*n),   "k0")
    row = _assumption_row(ws, row, "IT & systems",           f"{config['currency']}k","Input",
                          cst.get("it",           [400]*n),   "k0")
    row = _assumption_row(ws, row, "Professional fees",      f"{config['currency']}k","Input",
                          cst.get("prof_fees",    [350]*n),   "k0")
    row = _assumption_row(ws, row, "Other SG&A",             f"{config['currency']}k","Input",
                          cst.get("other_sga",   [500]*n),   "k0")
    row = _assumption_row(ws, row, "Restructuring / one-off",f"{config['currency']}k","Input",
                          cst.get("restr",        [0]*n),     "k0",
                          note="Non-recurring — excluded from adjusted EBITDA")
    row += 1

    # ────────────────────────────────────────────────────────────────────────
    # NWC DRIVERS
    # ────────────────────────────────────────────────────────────────────────
    row = _section_header(ws, row, "3. WORKING CAPITAL DRIVERS", "nwc_header", n + 3)
    nwc = config.get("nwc", {})
    row = _assumption_row(ws, row, "Days Sales Outstanding (DSO)",     "days","Input",
                          nwc.get("dso",  [45]*n), "int0")
    row = _assumption_row(ws, row, "Days Inventory Outstanding (DIO)", "days","Input",
                          nwc.get("dio",  [30]*n), "int0")
    row = _assumption_row(ws, row, "Days Payables Outstanding (DPO)",  "days","Input",
                          nwc.get("dpo",  [60]*n), "int0")
    row = _assumption_row(ws, row, "Other current assets (% revenue)", "%",  "Input",
                          nwc.get("oca_pct",[0.01]*n), "pct1")
    row = _assumption_row(ws, row, "Other current liabilities (% rev)","%" , "Input",
                          nwc.get("ocl_pct",[0.02]*n), "pct1")
    row += 1

    # ────────────────────────────────────────────────────────────────────────
    # CAPEX DRIVERS
    # ────────────────────────────────────────────────────────────────────────
    row = _section_header(ws, row, "4. CAPEX & DEPRECIATION DRIVERS", "kpi_header", n + 3)
    cpx = config.get("capex", {})
    row = _assumption_row(ws, row, "Maintenance capex",       f"{config['currency']}k","Input",
                          cpx.get("maint",  [3500]*n), "k0")
    row = _assumption_row(ws, row, "Expansion capex",         f"{config['currency']}k","Input",
                          cpx.get("expan",  [8000,5000,2000]+[1000]*(n-3) if n>3 else [5000]*n), "k0")
    row = _assumption_row(ws, row, "Opening PP&E (gross)",    f"{config['currency']}k","Input",
                          [cpx.get("opening_ppe", 85000)]+[None]*(n-1), "k0",
                          note="Year 1 only — subsequent years auto-calculated")
    row = _assumption_row(ws, row, "Depreciation — useful life (yrs)","yrs","Input",
                          [cpx.get("useful_life", 20)]+[None]*(n-1), "int0",
                          note="Straight-line — single rate for simplicity")
    row = _assumption_row(ws, row, "Accumulated depreciation (opening)",f"{config['currency']}k","Input",
                          [cpx.get("accum_dep_open", 24000)]+[None]*(n-1), "k0",
                          note="Year 1 only")
    row += 1

    # ────────────────────────────────────────────────────────────────────────
    # TAX DRIVERS
    # ────────────────────────────────────────────────────────────────────────
    row = _section_header(ws, row, "5. TAX DRIVERS", "kpi_header", n + 3)
    tax = config.get("tax", {})
    row = _assumption_row(ws, row, "Corporation tax rate",    "%",  "Input",
                          tax.get("rate",  [0.25]*n), "pct1")
    row = _assumption_row(ws, row, "Tax loss carry-forward (opening)",
                          f"{config['currency']}k","Input",
                          [tax.get("loss_cf_open",0)]+[None]*(n-1), "k0")
    row = _assumption_row(ws, row, "Deferred tax rate",       "%",  "Input",
                          tax.get("dt_rate",[0.25]*n), "pct1")
    row += 1

    # ────────────────────────────────────────────────────────────────────────
    # MACRO / FX
    # ────────────────────────────────────────────────────────────────────────
    row = _section_header(ws, row, "6. MACRO & FX ASSUMPTIONS", "kpi_header", n + 3)
    macro = config.get("macro", {})
    row = _assumption_row(ws, row, "CPI / inflation rate",    "%",  "Input",
                          macro.get("cpi",    [0.025]*n), "pct1")
    row = _assumption_row(ws, row, "USD/EUR FX rate",         "x",  "Input",
                          macro.get("fx_usd_eur",[1.09]*n), "k1")
    row = _assumption_row(ws, row, "USD/GBP FX rate",         "x",  "Input",
                          macro.get("fx_usd_gbp",[1.27]*n), "k1")
    row = _assumption_row(ws, row, "USD/AED FX rate",         "x",  "Input",
                          macro.get("fx_usd_aed",[3.67]*n), "k1")
    row += 1

    ws.sheet_properties.tabColor = "2E5F9E"
    return ws
